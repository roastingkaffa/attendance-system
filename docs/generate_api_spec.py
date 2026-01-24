#!/usr/bin/env python3
"""
生成 API 規格說明 Excel 檔案
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_api_spec_excel():
    wb = openpyxl.Workbook()

    # 樣式定義
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    category_font = Font(bold=True, size=11)
    category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    # ========== Sheet 1: API 總覽 ==========
    ws1 = wb.active
    ws1.title = "API總覽"

    headers = ["分類", "API名稱", "HTTP方法", "端點URL", "說明", "權限要求"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    api_list = [
        # 認證相關
        ("認證", "登入", "POST", "/login/", "使用帳號密碼登入系統", "無需認證"),
        ("認證", "登出", "POST", "/logout/", "登出系統", "無需認證"),
        ("認證", "修改密碼", "POST", "/change_password/", "修改當前使用者密碼", "需要認證"),
        ("認證", "忘記密碼", "POST", "/forgot_password/", "發送臨時密碼到 Email", "無需認證"),

        # 使用者資訊
        ("使用者", "取得個人資料", "GET", "/user/profile/", "取得當前使用者完整資訊（含角色、權限）", "需要認證"),
        ("使用者", "取得員工關聯", "GET", "/relation/?employee_id={id}", "取得員工與公司的關聯資訊", "需要認證"),

        # 打卡
        ("打卡", "上班打卡", "POST", "/clock-in/", "記錄上班打卡", "需要認證"),
        ("打卡", "下班打卡", "POST", "/clock-out/{record_id}/", "記錄下班打卡", "需要認證"),
        ("打卡", "查詢出勤記錄", "GET", "/attendance/?employee_id={id}&days={n}", "查詢指定天數內的出勤記錄", "需要認證"),

        # 請假
        ("請假", "申請請假", "POST", "/leave/apply/", "送出請假申請", "需要認證"),
        ("請假", "查詢請假記錄", "GET", "/leave/my-records/?days={n}&status={s}", "查詢個人請假記錄", "需要認證"),
        ("請假", "查詢假別額度", "GET", "/leave/balances/?year={y}", "查詢個人假別額度", "需要認證"),
        ("請假", "特休資格查詢", "GET", "/leave/annual-entitlement/", "查詢特休資格與天數", "需要認證"),
        ("請假", "計算特休天數", "POST", "/leave/calculate-annual/", "根據到職日計算特休", "需要認證"),

        # 加班
        ("加班", "申請加班", "POST", "/overtime/apply/", "送出加班申請", "需要認證"),
        ("加班", "查詢加班記錄", "GET", "/overtime/my-records/?days={n}", "查詢個人加班記錄", "需要認證"),
        ("加班", "取消加班申請", "POST", "/overtime/cancel/{overtime_id}/", "取消待審批的加班申請", "需要認證"),

        # 補打卡
        ("補打卡", "申請補打卡", "POST", "/makeup-clock/apply/", "送出補打卡申請", "需要認證"),
        ("補打卡", "查詢補打卡記錄", "GET", "/makeup-clock/my-requests/", "查詢個人補打卡申請", "需要認證"),
        ("補打卡", "查詢補打卡額度", "GET", "/makeup-clock/quota/", "查詢補打卡剩餘次數", "需要認證"),

        # 審批
        ("審批", "待審批清單", "GET", "/approval/pending/", "查詢待審批的申請", "需要認證"),
        ("審批", "請假待審批清單", "GET", "/leave/pending/", "查詢待審批的請假申請", "主管權限"),
        ("審批", "加班待審批清單", "GET", "/overtime/pending/", "查詢待審批的加班申請", "主管權限"),
        ("審批", "補打卡待審批清單", "GET", "/makeup-clock/pending/", "查詢待審批的補打卡申請", "主管權限"),
        ("審批", "批准請假", "POST", "/approval/approve/{approval_id}/", "批准請假申請", "主管權限"),
        ("審批", "拒絕請假", "POST", "/approval/reject/{approval_id}/", "拒絕請假申請", "主管權限"),
        ("審批", "批准加班", "POST", "/overtime/approve/{approval_id}/", "批准加班申請", "主管權限"),
        ("審批", "拒絕加班", "POST", "/overtime/reject/{approval_id}/", "拒絕加班申請", "主管權限"),
        ("審批", "批准補打卡", "POST", "/makeup-clock/approve/{approval_id}/", "批准補打卡申請", "主管權限"),
        ("審批", "拒絕補打卡", "POST", "/makeup-clock/reject/{approval_id}/", "拒絕補打卡申請", "主管權限"),
        ("審批", "批次審批", "POST", "/approval/batch/", "批次批准或拒絕多筆申請", "主管權限"),

        # 報表
        ("報表", "出勤摘要", "GET", "/reports/attendance-summary/?year={y}&month={m}", "查詢月度出勤摘要", "需要認證"),
        ("報表", "異常清單", "GET", "/reports/anomaly-list/?year={y}&month={m}", "查詢出勤異常清單", "需要認證"),

        # 通知
        ("通知", "查詢通知", "GET", "/notifications/?page={p}&page_size={s}", "查詢通知列表", "需要認證"),
        ("通知", "未讀通知數量", "GET", "/notifications/unread-count/", "查詢未讀通知數量", "需要認證"),
        ("通知", "標記已讀", "POST", "/notifications/mark-read/{notification_id}/", "標記單一通知為已讀", "需要認證"),
        ("通知", "全部標記已讀", "POST", "/notifications/mark-all-read/", "標記所有通知為已讀", "需要認證"),

        # 班表
        ("班表", "查詢班表", "GET", "/schedule/my-schedule/", "查詢個人班表設定", "需要認證"),

        # 主管功能
        ("主管", "部門出勤總覽", "GET", "/manager/dashboard/?date={d}", "查詢部門當日出勤狀況", "主管權限"),
        ("主管", "部門報表", "GET", "/manager/reports/department/?year={y}&month={m}", "查詢部門月度報表", "主管權限"),

        # HR 管理
        ("HR管理", "員工列表", "GET", "/hr/employees/?page={p}&search={s}&role={r}", "查詢員工列表（分頁、篩選）", "HR權限"),
        ("HR管理", "新增員工", "POST", "/hr/employees/create/", "新增員工資料", "HR權限"),
        ("HR管理", "更新員工", "PATCH", "/hr/employees/{employee_id}/", "更新員工資料", "HR權限"),
        ("HR管理", "指派主管", "PATCH", "/hr/employees/{employee_id}/assign-manager/", "為員工指派直屬主管", "HR權限"),
        ("HR管理", "批次設定假別額度", "POST", "/hr/leave-balances/batch-set/", "批次設定員工假別額度", "HR權限"),

        # 部門管理
        ("部門管理", "部門列表", "GET", "/hr/departments/?company_id={id}", "查詢部門列表", "HR權限"),
        ("部門管理", "新增部門", "POST", "/hr/departments/create/", "新增部門", "HR權限"),
        ("部門管理", "更新部門", "PATCH", "/hr/departments/{department_id}/", "更新部門資訊", "HR權限"),
        ("部門管理", "刪除部門", "DELETE", "/hr/departments/{department_id}/delete/", "刪除部門", "HR權限"),

        # 匯出
        ("匯出", "匯出出勤記錄", "POST", "/export/attendance/", "匯出出勤記錄為 CSV/Excel", "主管權限"),
        ("匯出", "匯出請假記錄", "POST", "/export/leave/", "匯出請假記錄為 CSV/Excel", "主管權限"),
    ]

    for row, api in enumerate(api_list, 2):
        for col, value in enumerate(api, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 45
    ws1.column_dimensions['E'].width = 35
    ws1.column_dimensions['F'].width = 12

    # ========== Sheet 2: 請求格式 ==========
    ws2 = wb.create_sheet("請求格式")

    headers2 = ["API名稱", "端點", "請求方法", "Content-Type", "請求參數", "參數類型", "必填", "說明"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    request_specs = [
        # 登入
        ("登入", "/login/", "POST", "application/json", "employee_id", "string", "是", "員工編號"),
        ("登入", "/login/", "POST", "application/json", "password", "string", "是", "密碼"),

        # 修改密碼
        ("修改密碼", "/change_password/", "POST", "application/json", "old_password", "string", "是", "舊密碼"),
        ("修改密碼", "/change_password/", "POST", "application/json", "new_password", "string", "是", "新密碼"),

        # 忘記密碼
        ("忘記密碼", "/forgot_password/", "POST", "application/json", "email", "string", "是", "註冊時使用的 Email"),

        # 打卡
        ("上班打卡", "/clock-in/", "POST", "application/json", "relation_id", "integer", "是", "員工-公司關聯 ID"),
        ("上班打卡", "/clock-in/", "POST", "application/json", "qr_data", "string", "是", "QR Code 資料"),
        ("上班打卡", "/clock-in/", "POST", "application/json", "latitude", "number", "是", "GPS 緯度"),
        ("上班打卡", "/clock-in/", "POST", "application/json", "longitude", "number", "是", "GPS 經度"),

        # 請假申請
        ("申請請假", "/leave/apply/", "POST", "application/json", "relation_id", "integer", "是", "員工-公司關聯 ID"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "leave_type", "string", "是", "假別: annual/sick/personal/funeral/marriage/maternity/paternity/menstrual/compensatory"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "start_time", "string", "是", "開始時間 (YYYY-MM-DD HH:MM:SS)"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "end_time", "string", "是", "結束時間 (YYYY-MM-DD HH:MM:SS)"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "leave_hours", "number", "是", "請假時數"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "leave_reason", "string", "是", "請假原因"),
        ("申請請假", "/leave/apply/", "POST", "application/json", "substitute_employee_id", "string", "否", "職務代理人員工編號"),

        # 加班申請
        ("申請加班", "/overtime/apply/", "POST", "application/json", "date", "string", "是", "加班日期 (YYYY-MM-DD)"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "start_time", "string", "是", "開始時間 (HH:MM)"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "end_time", "string", "是", "結束時間 (HH:MM)"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "reason", "string", "是", "加班原因"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "compensation_type", "string", "否", "補償方式: pay/compensatory/mixed，預設 compensatory"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "compensatory_hours", "number", "否", "補休時數（mixed 時使用）"),
        ("申請加班", "/overtime/apply/", "POST", "application/json", "pay_hours", "number", "否", "加班費時數（mixed 時使用）"),

        # 補打卡申請
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "relation_id", "integer", "是", "員工-公司關聯 ID"),
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "date", "string", "是", "補打卡日期 (YYYY-MM-DD)"),
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "makeup_type", "string", "是", "類型: checkin/checkout/both"),
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "requested_checkin_time", "string", "否", "申請的上班時間 (HH:MM)"),
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "requested_checkout_time", "string", "否", "申請的下班時間 (HH:MM)"),
        ("申請補打卡", "/makeup-clock/apply/", "POST", "application/json", "reason", "string", "是", "補打卡原因"),

        # 審批
        ("批准/拒絕申請", "/approval/approve/{id}/", "POST", "application/json", "comment", "string", "否", "審批意見"),

        # 批次審批
        ("批次審批", "/approval/batch/", "POST", "application/json", "approval_type", "string", "是", "審批類型: leave/overtime/makeup"),
        ("批次審批", "/approval/batch/", "POST", "application/json", "approval_ids", "array", "是", "審批記錄 ID 陣列 [1,2,3]"),
        ("批次審批", "/approval/batch/", "POST", "application/json", "action", "string", "是", "動作: approve/reject"),
        ("批次審批", "/approval/batch/", "POST", "application/json", "comment", "string", "否", "審批意見"),

        # 新增員工
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "employee_id", "string", "是", "員工編號（唯一）"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "username", "string", "是", "姓名"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "email", "string", "是", "Email"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "phone", "string", "否", "電話"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "role", "string", "否", "角色: employee/manager/hr_admin/ceo/system_admin"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "department_id", "integer", "否", "部門 ID"),
        ("新增員工", "/hr/employees/create/", "POST", "application/json", "password", "string", "否", "初始密碼（未提供則自動生成）"),

        # 匯出
        ("匯出出勤記錄", "/export/attendance/", "POST", "application/json", "date_from", "string", "是", "開始日期 (YYYY-MM-DD)"),
        ("匯出出勤記錄", "/export/attendance/", "POST", "application/json", "date_to", "string", "是", "結束日期 (YYYY-MM-DD)"),
        ("匯出出勤記錄", "/export/attendance/", "POST", "application/json", "format", "string", "否", "格式: csv/xlsx，預設 csv"),
        ("匯出出勤記錄", "/export/attendance/", "POST", "application/json", "employee_ids", "array", "否", "員工編號陣列（HR 專用）"),
    ]

    for row, spec in enumerate(request_specs, 2):
        for col, value in enumerate(spec, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 8
    ws2.column_dimensions['H'].width = 45

    # ========== Sheet 3: 回應格式 ==========
    ws3 = wb.create_sheet("回應格式")

    headers3 = ["分類", "欄位路徑", "資料類型", "說明", "範例值"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    response_specs = [
        # 標準回應格式
        ("標準成功回應", "success", "boolean", "是否成功", "true"),
        ("標準成功回應", "message", "string", "回應訊息", "操作成功"),
        ("標準成功回應", "data", "object/array", "回應資料", "{}"),

        ("標準錯誤回應", "success", "boolean", "是否成功", "false"),
        ("標準錯誤回應", "error.message", "string", "錯誤訊息", "缺少必要參數"),
        ("標準錯誤回應", "error.code", "string", "錯誤代碼", "VALIDATION_ERROR"),
        ("標準錯誤回應", "error.details", "object", "錯誤詳情", "null"),

        # 登入回應
        ("登入回應", "data.userId", "string", "員工編號", "EMP001"),

        # 使用者資訊
        ("使用者資訊", "data.employee_id", "string", "員工編號", "EMP001"),
        ("使用者資訊", "data.username", "string", "姓名", "張小明"),
        ("使用者資訊", "data.email", "string", "Email", "emp001@example.com"),
        ("使用者資訊", "data.phone", "string", "電話", "0912345678"),
        ("使用者資訊", "data.role", "string", "角色代碼", "employee"),
        ("使用者資訊", "data.role_display", "string", "角色顯示名稱", "一般員工"),
        ("使用者資訊", "data.department", "integer", "部門 ID", "1"),
        ("使用者資訊", "data.department_name", "string", "部門名稱", "研發部"),
        ("使用者資訊", "data.permissions", "object", "權限物件", "{view_own_attendance: true, ...}"),
        ("使用者資訊", "data.company_info", "array", "公司關聯資訊", "[{relation_id: 1, company_name: '...'}]"),

        # 出勤記錄
        ("出勤記錄", "[].id", "integer", "記錄 ID", "1"),
        ("出勤記錄", "[].date", "string", "日期", "2026-01-24"),
        ("出勤記錄", "[].checkin_time", "string", "上班時間", "2026-01-24T09:00:00"),
        ("出勤記錄", "[].checkout_time", "string", "下班時間", "2026-01-24T18:00:00"),
        ("出勤記錄", "[].work_hours", "number", "工時", "8.0"),
        ("出勤記錄", "[].is_late", "boolean", "是否遲到", "false"),
        ("出勤記錄", "[].late_minutes", "integer", "遲到分鐘數", "0"),
        ("出勤記錄", "[].is_early_leave", "boolean", "是否早退", "false"),
        ("出勤記錄", "[].early_leave_minutes", "integer", "早退分鐘數", "0"),
        ("出勤記錄", "[].is_makeup", "boolean", "是否補打卡", "false"),

        # 請假記錄
        ("請假記錄", "data.records[].id", "integer", "記錄 ID", "1"),
        ("請假記錄", "data.records[].leave_type", "string", "假別代碼", "annual"),
        ("請假記錄", "data.records[].leave_type_display", "string", "假別名稱", "特休"),
        ("請假記錄", "data.records[].start_time", "string", "開始時間", "2026-01-24T09:00:00"),
        ("請假記錄", "data.records[].end_time", "string", "結束時間", "2026-01-24T18:00:00"),
        ("請假記錄", "data.records[].leave_hours", "number", "請假時數", "8.0"),
        ("請假記錄", "data.records[].status", "string", "狀態代碼", "pending"),
        ("請假記錄", "data.records[].status_display", "string", "狀態名稱", "待審批"),
        ("請假記錄", "data.records[].leave_reason", "string", "請假原因", "家庭事務"),

        # 假別額度
        ("假別額度", "data.year", "integer", "年度", "2026"),
        ("假別額度", "data.summary.total_hours", "number", "總額度時數", "120.0"),
        ("假別額度", "data.summary.used_hours", "number", "已使用時數", "16.0"),
        ("假別額度", "data.summary.remaining_hours", "number", "剩餘時數", "104.0"),
        ("假別額度", "data.balances[].leave_type", "string", "假別代碼", "annual"),
        ("假別額度", "data.balances[].leave_type_display", "string", "假別名稱", "特休"),
        ("假別額度", "data.balances[].total_hours", "number", "總額度", "80.0"),
        ("假別額度", "data.balances[].used_hours", "number", "已使用", "8.0"),
        ("假別額度", "data.balances[].remaining_hours", "number", "剩餘", "72.0"),

        # 加班記錄
        ("加班記錄", "data.records[].id", "integer", "記錄 ID", "1"),
        ("加班記錄", "data.records[].date", "string", "加班日期", "2026-01-24"),
        ("加班記錄", "data.records[].start_time", "string", "開始時間", "18:00:00"),
        ("加班記錄", "data.records[].end_time", "string", "結束時間", "21:00:00"),
        ("加班記錄", "data.records[].overtime_hours", "number", "加班時數", "3.0"),
        ("加班記錄", "data.records[].compensation_type", "string", "補償方式", "compensatory"),
        ("加班記錄", "data.records[].compensation_display", "string", "補償方式名稱", "補休"),
        ("加班記錄", "data.records[].status", "string", "狀態代碼", "pending"),
        ("加班記錄", "data.records[].status_display", "string", "狀態名稱", "待審批"),

        # 補打卡額度
        ("補打卡額度", "data.year", "integer", "年度", "2026"),
        ("補打卡額度", "data.total_count", "integer", "總次數", "24"),
        ("補打卡額度", "data.used_count", "integer", "已使用次數", "2"),
        ("補打卡額度", "data.remaining_count", "integer", "剩餘次數", "22"),

        # 通知
        ("通知", "data.unread_count", "integer", "未讀數量", "5"),
        ("通知", "data.notifications[].id", "integer", "通知 ID", "1"),
        ("通知", "data.notifications[].title", "string", "標題", "請假申請已批准"),
        ("通知", "data.notifications[].message", "string", "內容", "您的請假申請已獲批准"),
        ("通知", "data.notifications[].type", "string", "類型", "approval"),
        ("通知", "data.notifications[].is_read", "boolean", "是否已讀", "false"),
        ("通知", "data.notifications[].created_at", "string", "建立時間", "2026-01-24T10:00:00"),

        # HR 員工列表
        ("HR員工列表", "data.total", "integer", "總筆數", "50"),
        ("HR員工列表", "data.page", "integer", "當前頁碼", "1"),
        ("HR員工列表", "data.page_size", "integer", "每頁筆數", "20"),
        ("HR員工列表", "data.total_pages", "integer", "總頁數", "3"),
        ("HR員工列表", "data.employees[].employee_id", "string", "員工編號", "EMP001"),
        ("HR員工列表", "data.employees[].username", "string", "姓名", "張小明"),
        ("HR員工列表", "data.employees[].role", "string", "角色代碼", "employee"),
        ("HR員工列表", "data.employees[].role_display", "string", "角色名稱", "一般員工"),
        ("HR員工列表", "data.employees[].is_active", "boolean", "是否在職", "true"),
    ]

    for row, spec in enumerate(response_specs, 2):
        for col, value in enumerate(spec, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 30
    ws3.column_dimensions['E'].width = 40

    # ========== Sheet 4: 錯誤代碼 ==========
    ws4 = wb.create_sheet("錯誤代碼")

    headers4 = ["HTTP狀態碼", "錯誤代碼", "說明", "可能原因", "建議處理方式"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    error_codes = [
        ("400", "VALIDATION_ERROR", "參數驗證失敗", "缺少必填參數或格式錯誤", "檢查請求參數是否完整且格式正確"),
        ("400", "MISSING_PARAMETERS", "缺少必要參數", "請求未包含必填欄位", "確認所有必填欄位已填入"),
        ("400", "INVALID_DATE", "日期無效", "日期格式錯誤或超出允許範圍", "確認日期格式為 YYYY-MM-DD"),
        ("400", "INSUFFICIENT_BALANCE", "額度不足", "假別剩餘時數不足", "檢查假別額度或減少請假時數"),
        ("400", "QUOTA_EXCEEDED", "額度已用完", "補打卡次數已達上限", "聯繫 HR 調整額度"),
        ("400", "DUPLICATE_REQUEST", "重複申請", "該時段已有申請記錄", "取消或修改既有申請"),
        ("401", "INVALID_CREDENTIALS", "認證失敗", "帳號或密碼錯誤", "確認帳號密碼是否正確"),
        ("401", "UNAUTHORIZED", "未授權", "未登入或 Session 過期", "重新登入系統"),
        ("403", "FORBIDDEN", "權限不足", "無權執行此操作", "確認帳號角色是否有權限"),
        ("403", "PERMISSION_DENIED", "權限拒絕", "無主管或 HR 權限", "聯繫系統管理員"),
        ("404", "NOT_FOUND", "資源不存在", "查詢的記錄不存在", "確認 ID 是否正確"),
        ("404", "INVALID_RELATION", "無效的關聯", "員工-公司關聯不存在", "確認 relation_id 是否正確"),
        ("500", "INTERNAL_ERROR", "伺服器錯誤", "系統內部錯誤", "稍後再試或聯繫系統管理員"),
    ]

    for row, error in enumerate(error_codes, 2):
        for col, value in enumerate(error, 1):
            cell = ws4.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 30
    ws4.column_dimensions['E'].width = 35

    # ========== Sheet 5: 資料類型說明 ==========
    ws5 = wb.create_sheet("資料類型說明")

    headers5 = ["分類", "代碼", "說明"]
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_types = [
        # 角色
        ("角色 (role)", "employee", "一般員工"),
        ("角色 (role)", "manager", "部門主管"),
        ("角色 (role)", "hr_admin", "HR 管理員"),
        ("角色 (role)", "ceo", "總經理"),
        ("角色 (role)", "system_admin", "系統管理員"),

        # 假別
        ("假別 (leave_type)", "annual", "特休"),
        ("假別 (leave_type)", "sick", "病假"),
        ("假別 (leave_type)", "personal", "事假"),
        ("假別 (leave_type)", "funeral", "喪假"),
        ("假別 (leave_type)", "marriage", "婚假"),
        ("假別 (leave_type)", "maternity", "產假"),
        ("假別 (leave_type)", "paternity", "陪產假"),
        ("假別 (leave_type)", "menstrual", "生理假"),
        ("假別 (leave_type)", "compensatory", "補休"),
        ("假別 (leave_type)", "official", "公假"),

        # 審批狀態
        ("審批狀態 (status)", "pending", "待審批"),
        ("審批狀態 (status)", "approved", "已批准"),
        ("審批狀態 (status)", "rejected", "已拒絕"),
        ("審批狀態 (status)", "cancelled", "已取消"),

        # 補打卡類型
        ("補打卡類型 (makeup_type)", "checkin", "上班打卡"),
        ("補打卡類型 (makeup_type)", "checkout", "下班打卡"),
        ("補打卡類型 (makeup_type)", "both", "上下班都補"),

        # 加班補償方式
        ("補償方式 (compensation_type)", "pay", "加班費"),
        ("補償方式 (compensation_type)", "compensatory", "補休"),
        ("補償方式 (compensation_type)", "mixed", "混合（部分補休、部分加班費）"),

        # 通知類型
        ("通知類型 (notification_type)", "approval", "審批結果通知"),
        ("通知類型 (notification_type)", "system", "系統通知"),
        ("通知類型 (notification_type)", "reminder", "提醒通知"),

        # 匯出格式
        ("匯出格式 (format)", "csv", "CSV 格式（逗號分隔）"),
        ("匯出格式 (format)", "xlsx", "Excel 格式"),
    ]

    for row, item in enumerate(data_types, 2):
        for col, value in enumerate(item, 1):
            cell = ws5.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 40

    # ========== Sheet 6: 範例 JSON ==========
    ws6 = wb.create_sheet("範例JSON")

    headers6 = ["API", "類型", "JSON 範例"]
    for col, header in enumerate(headers6, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    json_examples = [
        ("登入", "Request", '{\n  "employee_id": "EMP001",\n  "password": "password123"\n}'),
        ("登入", "Response (成功)", '{\n  "success": true,\n  "message": "登入成功",\n  "data": {\n    "userId": "EMP001"\n  }\n}'),
        ("登入", "Response (失敗)", '{\n  "success": false,\n  "error": {\n    "message": "登入失敗，請檢查帳號密碼",\n    "code": "INVALID_CREDENTIALS",\n    "details": null\n  }\n}'),

        ("申請請假", "Request", '{\n  "relation_id": 1,\n  "leave_type": "annual",\n  "start_time": "2026-01-25 09:00:00",\n  "end_time": "2026-01-25 18:00:00",\n  "leave_hours": 8.0,\n  "leave_reason": "家庭事務",\n  "substitute_employee_id": "EMP002"\n}'),
        ("申請請假", "Response", '{\n  "success": true,\n  "message": "請假申請已送出",\n  "data": {\n    "leave_id": 123,\n    "approval_id": 456,\n    "status": "pending"\n  }\n}'),

        ("申請加班", "Request", '{\n  "date": "2026-01-24",\n  "start_time": "18:00",\n  "end_time": "21:00",\n  "reason": "專案趕工",\n  "compensation_type": "compensatory"\n}'),
        ("申請加班", "Response", '{\n  "success": true,\n  "message": "加班申請已送出",\n  "data": {\n    "overtime_id": 789,\n    "overtime_hours": 3.0,\n    "status": "pending"\n  }\n}'),

        ("申請補打卡", "Request", '{\n  "relation_id": 1,\n  "date": "2026-01-23",\n  "makeup_type": "checkin",\n  "requested_checkin_time": "09:00",\n  "reason": "忘記打卡"\n}'),
        ("申請補打卡", "Response", '{\n  "success": true,\n  "message": "補打卡申請已送出",\n  "data": {\n    "request_id": 101,\n    "status": "pending"\n  }\n}'),

        ("批次審批", "Request", '{\n  "approval_type": "leave",\n  "approval_ids": [1, 2, 3],\n  "action": "approve",\n  "comment": "核准"\n}'),
        ("批次審批", "Response", '{\n  "success": true,\n  "message": "批次審批完成",\n  "data": {\n    "processed": 3,\n    "success_count": 3,\n    "failed_count": 0\n  }\n}'),

        ("取得使用者資訊", "Response", '{\n  "success": true,\n  "message": "查詢成功",\n  "data": {\n    "employee_id": "EMP001",\n    "username": "張小明",\n    "email": "emp001@example.com",\n    "role": "employee",\n    "role_display": "一般員工",\n    "department_name": "研發部",\n    "permissions": {\n      "view_own_attendance": true,\n      "apply_leave": true,\n      "apply_overtime": true\n    }\n  }\n}'),

        ("假別額度查詢", "Response", '{\n  "success": true,\n  "message": "查詢成功",\n  "data": {\n    "year": 2026,\n    "summary": {\n      "total_hours": 120.0,\n      "used_hours": 16.0,\n      "remaining_hours": 104.0\n    },\n    "balances": [\n      {\n        "leave_type": "annual",\n        "leave_type_display": "特休",\n        "total_hours": 80.0,\n        "used_hours": 8.0,\n        "remaining_hours": 72.0\n      }\n    ]\n  }\n}'),

        ("匯出出勤記錄", "Request", '{\n  "date_from": "2026-01-01",\n  "date_to": "2026-01-31",\n  "format": "xlsx",\n  "employee_ids": ["EMP001", "EMP002"]\n}'),
        ("匯出出勤記錄", "Response", "Binary file (Excel/CSV) 直接下載"),
    ]

    for row, example in enumerate(json_examples, 2):
        for col, value in enumerate(example, 1):
            cell = ws6.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 設定欄寬與高度
    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 15
    ws6.column_dimensions['C'].width = 80

    # 設定行高
    for row in range(2, len(json_examples) + 2):
        ws6.row_dimensions[row].height = 120

    # 儲存檔案
    output_path = "/home/roc/workspace/Human-Resources/attendance-system/docs/API_SPECIFICATION.xlsx"
    wb.save(output_path)
    print(f"Excel 檔案已建立: {output_path}")
    return output_path

if __name__ == "__main__":
    create_api_spec_excel()
