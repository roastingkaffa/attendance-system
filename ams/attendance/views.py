from django.contrib.auth import authenticate, get_user_model, login as django_login, logout
from django.core.mail import send_mail
from django.http import HttpResponse
from django.db.models import Q
from .models import *
from .serializers import *
from .responses import success_response, error_response, unauthorized_response, validation_error_response, server_error_response
from datetime import time, date, timedelta, datetime
from rest_framework.viewsets import ModelViewSet
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework.decorators import permission_classes
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.decorators import api_view, permission_classes, authentication_classes
import random, string


@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([SessionAuthentication])
def login(request):
    userId = request.data.get("userId")
    password = request.data.get("password")

    if not userId or not password:
        return validation_error_response("請提供帳號和密碼")

    user = authenticate(username=userId, password=password)
    if user is not None:
        django_login(request._request, user)
        return success_response(
            message="登入成功",
            data={"userId": userId}
        )
    else:
        return unauthorized_response("登入失敗，請檢查帳號密碼", code="INVALID_CREDENTIALS")


@api_view(['POST'])
@permission_classes([IsAuthenticated])  # 🔒 修正：改為 IsAuthenticated
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not old_password or not new_password:
        return validation_error_response("請提供舊密碼和新密碼")

    if not user.check_password(old_password):
        return error_response("舊密碼不正確", code="INVALID_OLD_PASSWORD")

    user.set_password(new_password)
    user.save()
    return success_response(message="密碼已更新")


@api_view(["POST"])
@permission_classes([AllowAny])
def logout_user(request):
    logout(request)
    return success_response(message="登出成功")


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    email = request.data.get('email')

    if not email:
        return validation_error_response("請提供 Email")

    email = email.strip()

    user = Employees.objects.filter(email__iexact=email).first()
    if not user:
        # 安全考量：即使找不到使用者也回傳成功訊息，避免洩露使用者是否存在
        print(f"⚠️ 找不到使用者：{email}")
        return success_response(message="如果該 Email 存在，臨時密碼已寄出")

    print(f"✅ 找到使用者：{user.email}")

    temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    user.set_password(temp_password)
    user.save()

    try:
        send_mail(
            subject='臨時密碼通知',
            message=f'您的臨時密碼為：{temp_password}\n請使用此密碼登入並盡快修改。',
            from_email='4104W008@gmail.com',
            recipient_list=[email],
            fail_silently=False
        )
        print("✅ 郵件成功寄出")
        return success_response(message="臨時密碼已寄出")
    except Exception as e:
        print(f"❌ 發信錯誤：{e}")
        return server_error_response("發送郵件失敗，請稍後再試", code="EMAIL_SEND_FAILED")


def test_send_email(request):
    try:
        send_mail(
            subject='測試信',
            message='這是一封測試郵件',
            from_email='sax0224@gmail.com',
            recipient_list=['sax0224@yahoo.com'],
            fail_silently=False
        )
        return HttpResponse("✅ 信件已成功寄出")
    except Exception as e:
        return HttpResponse(f"❌ 發送失敗：{str(e)}")

@permission_classes((IsAuthenticated,))
class EmployeesView(ModelViewSet):
    authentication_classes = (SessionAuthentication, BasicAuthentication)
    permission_classes = [IsAuthenticated]
    queryset = Employees.objects.all()
    serializer_class = EmployeesSerializer

    def create(self, request, *args, **kwargs):
        # 判斷是不是傳 list，若是就啟用 many=True
        is_many = isinstance(request.data, list)

        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        serializer.save()

@permission_classes((IsAuthenticated,))
class CompaniesView(ModelViewSet):
    authentication_classes = (SessionAuthentication, BasicAuthentication)
    permission_classes = [IsAuthenticated]
    queryset = Companies.objects.all()
    serializer_class = CompaniesSerializer


@permission_classes((IsAuthenticated,))
class leaveRecordsView(viewsets.ModelViewSet):
    authentication_classes = (SessionAuthentication, BasicAuthentication)
    permission_classes = [IsAuthenticated]
    queryset = LeaveRecords.objects.all()
    serializer_class = LeaveRecordsSerializer

    def get_queryset(self):
        queryset = LeaveRecords.objects.all()

        employee_id = self.request.query_params.get("employee_id")
        days = self.request.query_params.get("days")

        if employee_id:
            queryset = queryset.filter(relation_id__employee_id__employee_id=employee_id)

        if days:
            days = int(days)
            start_date = date.today() - timedelta(days=days)
            end_date = date.today()
            end_date = datetime.combine(end_date, time.max)
            queryset = queryset.filter(start_time__range=[start_date, end_date])

        return queryset

@permission_classes((IsAuthenticated,))
class AttendanceRecordsView(viewsets.ModelViewSet):
    authentication_classes = (SessionAuthentication, BasicAuthentication)
    permission_classes = [IsAuthenticated]
    queryset = AttendanceRecords.objects.all()
    serializer_class = AttendanceRecordsSerializer

    def get_queryset(self):
        queryset = AttendanceRecords.objects.all()

        employee_id = self.request.query_params.get("employee_id")
        days = self.request.query_params.get("days")
        date_parm = self.request.query_params.get("date")

        if employee_id:
            queryset = queryset.filter(relation_id__employee_id__employee_id=employee_id)

        if days:
            days = int(days)
            start_date = date.today() - timedelta(days=days)
            end_date = date.today()
            queryset = queryset.filter(date__range=[start_date, end_date])

        if date_parm == "today":
            today = date.today()
            queryset = queryset.filter(date=today)

        return queryset

@permission_classes((IsAuthenticated,))
class EmpCompanyRelView(viewsets.ModelViewSet):
    authentication_classes = (SessionAuthentication, BasicAuthentication)
    permission_classes = [IsAuthenticated]
    queryset = EmpCompanyRel.objects.all()
    serializer_class = EmpCompanyRelSerializer

    def get_queryset(self):
        queryset = EmpCompanyRel.objects.all()
        employee_id = self.request.query_params.get("employee_id")
        if employee_id:
            queryset = queryset.filter(employee_id__employee_id=employee_id)
        return queryset


# ========== 新增：後端打卡驗證 API ==========
from .utils import calculate_distance, calculate_work_hours
from django.utils import timezone
from decimal import Decimal


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clock_in(request):
    """
    上班打卡 API（後端驗證版本）

    請求參數：
    - qr_latitude: QR Code 緯度
    - qr_longitude: QR Code 經度
    - user_latitude: 使用者緯度
    - user_longitude: 使用者經度
    - relation_id: 員工-公司關聯 ID
    """
    try:
        # 1. 取得請求參數
        qr_lat = request.data.get('qr_latitude')
        qr_lng = request.data.get('qr_longitude')
        user_lat = request.data.get('user_latitude')
        user_lng = request.data.get('user_longitude')
        relation_id = request.data.get('relation_id')

        # 2. 參數驗證
        if not all([qr_lat, qr_lng, user_lat, user_lng, relation_id]):
            return Response({
                'success': False,
                'error': {
                    'code': 'MISSING_PARAMETERS',
                    'message': '缺少必要參數'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 3. 轉換參數類型
        try:
            qr_lat = Decimal(str(qr_lat))
            qr_lng = Decimal(str(qr_lng))
            user_lat = Decimal(str(user_lat))
            user_lng = Decimal(str(user_lng))
        except Exception:
            return Response({
                'success': False,
                'error': {
                    'code': 'INVALID_COORDINATES',
                    'message': 'GPS 座標格式錯誤'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 4. 驗證 QR Code 座標是否為有效公司
        company = Companies.objects.filter(
            latitude=str(qr_lat),
            longitude=str(qr_lng)
        ).first()

        if not company:
            return Response({
                'success': False,
                'error': {
                    'code': 'INVALID_QR_CODE',
                    'message': '無效的 QR Code'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 5. 計算 GPS 距離（後端計算）
        distance = calculate_distance(user_lat, user_lng, qr_lat, qr_lng)

        # 6. 驗證距離是否在範圍內（預設 2000 公尺）
        max_distance = float(company.radius) if company.radius else 2000.0
        if distance > max_distance:
            return Response({
                'success': False,
                'error': {
                    'code': 'LOCATION_OUT_OF_RANGE',
                    'message': f'打卡位置超出範圍（{max_distance} 公尺）',
                    'details': {
                        'distance': round(distance, 2),
                        'max_distance': max_distance
                    }
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 7. 檢查今天是否已經打卡
        today = timezone.now().date()
        existing_record = AttendanceRecords.objects.filter(
            relation_id=relation_id,
            date=today
        ).first()

        if existing_record:
            return Response({
                'success': False,
                'error': {
                    'code': 'ALREADY_CLOCKED_IN',
                    'message': '今天已經打過卡'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 8. 產生當前時間（後端產生）
        now = timezone.now()
        location = f"{user_lat}, {user_lng}"

        # =====================================================
        # Phase 1 新增：取得員工班表並判定遲到
        # =====================================================
        from .models import WorkSchedule, EmpCompanyRel
        from datetime import datetime, timedelta

        # 取得員工關聯資訊
        try:
            relation = EmpCompanyRel.objects.get(id=relation_id)
        except EmpCompanyRel.DoesNotExist:
            return Response({
                'success': False,
                'error': {
                    'code': 'RELATION_NOT_FOUND',
                    'message': '無效的員工關聯'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 取得班表（優先員工專屬，否則使用公司預設）
        schedule = relation.work_schedule
        if not schedule:
            schedule = WorkSchedule.objects.filter(
                company_id=relation.company_id,
                is_default=True,
                is_active=True
            ).first()

        # 判定遲到
        is_late = False
        late_minutes = 0

        if schedule:
            # 組合今日的上班時間
            scheduled_start = timezone.make_aware(
                datetime.combine(today, schedule.work_start_time)
            )
            # 加上寬限時間
            grace_deadline = scheduled_start + timedelta(minutes=schedule.grace_period_minutes)

            if now > grace_deadline:
                is_late = True
                late_minutes = int((now - scheduled_start).total_seconds() / 60)
                if late_minutes < 0:
                    late_minutes = 0

        # 9. 建立打卡記錄（含遲到資訊）
        record = AttendanceRecords.objects.create(
            relation_id_id=relation_id,
            date=today,
            checkin_time=now,
            checkout_time=now,  # 初始設定為相同時間
            checkin_location=location,
            checkout_location=location,
            work_hours=Decimal('0.00'),
            schedule=schedule,      # Phase 1 新增
            is_late=is_late,        # Phase 1 新增
            late_minutes=late_minutes  # Phase 1 新增
        )

        # 10. 返回成功回應（含遲到資訊）
        response_data = {
            'id': record.id,
            'date': str(record.date),
            'checkin_time': record.checkin_time.isoformat(),
            'checkin_location': record.checkin_location,
            'distance': round(distance, 2)
        }

        # Phase 1 新增：遲到提示
        if is_late:
            response_data['is_late'] = True
            response_data['late_minutes'] = late_minutes
            message = f'打卡成功，但您已遲到 {late_minutes} 分鐘'
        else:
            response_data['is_late'] = False
            response_data['late_minutes'] = 0
            message = '打卡成功'

        return Response({
            'success': True,
            'message': message,
            'data': response_data
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        print(f"打卡錯誤: {str(e)}")
        return Response({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': '打卡失敗，請稍後再試'
            }
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def clock_out(request, record_id):
    """
    下班打卡 API（後端驗證版本）

    請求參數：
    - qr_latitude: QR Code 緯度
    - qr_longitude: QR Code 經度
    - user_latitude: 使用者緯度
    - user_longitude: 使用者經度
    """
    try:
        # 1. 取得打卡記錄
        try:
            record = AttendanceRecords.objects.get(id=record_id)
        except AttendanceRecords.DoesNotExist:
            return Response({
                'success': False,
                'error': {
                    'code': 'RECORD_NOT_FOUND',
                    'message': '找不到打卡記錄'
                }
            }, status=status.HTTP_404_NOT_FOUND)

        # 2. 取得請求參數
        qr_lat = request.data.get('qr_latitude')
        qr_lng = request.data.get('qr_longitude')
        user_lat = request.data.get('user_latitude')
        user_lng = request.data.get('user_longitude')

        # 3. 參數驗證
        if not all([qr_lat, qr_lng, user_lat, user_lng]):
            return Response({
                'success': False,
                'error': {
                    'code': 'MISSING_PARAMETERS',
                    'message': '缺少必要參數'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 4. 轉換參數類型
        try:
            qr_lat = Decimal(str(qr_lat))
            qr_lng = Decimal(str(qr_lng))
            user_lat = Decimal(str(user_lat))
            user_lng = Decimal(str(user_lng))
        except Exception:
            return Response({
                'success': False,
                'error': {
                    'code': 'INVALID_COORDINATES',
                    'message': 'GPS 座標格式錯誤'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 5. 驗證 QR Code
        company = Companies.objects.filter(
            latitude=str(qr_lat),
            longitude=str(qr_lng)
        ).first()

        if not company:
            return Response({
                'success': False,
                'error': {
                    'code': 'INVALID_QR_CODE',
                    'message': '無效的 QR Code'
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 6. 計算距離
        distance = calculate_distance(user_lat, user_lng, qr_lat, qr_lng)
        max_distance = float(company.radius) if company.radius else 2000.0

        if distance > max_distance:
            return Response({
                'success': False,
                'error': {
                    'code': 'LOCATION_OUT_OF_RANGE',
                    'message': f'打卡位置超出範圍（{max_distance} 公尺）',
                    'details': {
                        'distance': round(distance, 2),
                        'max_distance': max_distance
                    }
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        # 7. 產生當前時間（後端產生）
        now = timezone.now()
        location = f"{user_lat}, {user_lng}"

        # 8. 計算工時（後端計算）
        work_hours = calculate_work_hours(record.checkin_time, now)

        # =====================================================
        # Phase 1 新增：判定早退
        # =====================================================
        from datetime import datetime, timedelta

        is_early_leave = False
        early_leave_minutes = 0

        schedule = record.schedule
        if schedule:
            # 組合今日的下班時間
            scheduled_end = timezone.make_aware(
                datetime.combine(record.date, schedule.work_end_time)
            )

            if now < scheduled_end:
                is_early_leave = True
                early_leave_minutes = int((scheduled_end - now).total_seconds() / 60)
                if early_leave_minutes < 0:
                    early_leave_minutes = 0

        # 9. 更新記錄（含早退資訊）
        record.checkout_time = now
        record.checkout_location = location
        record.work_hours = work_hours
        record.is_early_leave = is_early_leave      # Phase 1 新增
        record.early_leave_minutes = early_leave_minutes  # Phase 1 新增
        record.save()

        # 10. 返回成功回應（含早退資訊）
        response_data = {
            'id': record.id,
            'date': str(record.date),
            'checkin_time': record.checkin_time.isoformat(),
            'checkout_time': record.checkout_time.isoformat(),
            'work_hours': float(record.work_hours),
            'distance': round(distance, 2)
        }

        # Phase 1 新增：早退提示
        if is_early_leave:
            response_data['is_early_leave'] = True
            response_data['early_leave_minutes'] = early_leave_minutes
            message = f'打卡成功，但您提早 {early_leave_minutes} 分鐘下班'
        else:
            response_data['is_early_leave'] = False
            response_data['early_leave_minutes'] = 0
            message = '打卡成功'

        return Response({
            'success': True,
            'message': message,
            'data': response_data
        }, status=status.HTTP_200_OK)

    except Exception as e:
        print(f"下班打卡錯誤: {str(e)}")
        return Response({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': '打卡失敗，請稍後再試'
            }
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ========== Phase 2 Week 4: 請假與審批 API ==========
from .models import ApprovalRecords, LeaveBalances
from .serializers import LeaveRecordsSerializer, ApprovalRecordsSerializer, LeaveBalancesSerializer
from decimal import Decimal as D


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_leave(request):
    """
    請假申請 API（Phase 2 Week 4）

    請求參數：
    - relation_id: 員工-公司關聯 ID
    - leave_type: 假別（annual, sick, personal, etc.）
    - start_time: 請假開始時間（YYYY-MM-DD HH:MM:SS）
    - end_time: 請假結束時間（YYYY-MM-DD HH:MM:SS）
    - leave_hours: 請假時數
    - leave_reason: 請假原因
    - substitute_employee_id: 職務代理人（選填）
    """
    try:
        # 1. 取得請求參數
        relation_id = request.data.get('relation_id')
        leave_type = request.data.get('leave_type')
        start_time = request.data.get('start_time')
        end_time = request.data.get('end_time')
        leave_hours = request.data.get('leave_hours')
        leave_reason = request.data.get('leave_reason', '')
        substitute_employee_id = request.data.get('substitute_employee_id')

        # 2. 參數驗證
        if not all([relation_id, leave_type, start_time, end_time, leave_hours]):
            return error_response(
                "缺少必要參數",
                code="MISSING_PARAMETERS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 3. 驗證 relation_id
        try:
            relation = EmpCompanyRel.objects.get(id=relation_id)
            employee = relation.employee_id
        except EmpCompanyRel.DoesNotExist:
            return error_response(
                "無效的員工-公司關聯",
                code="INVALID_RELATION",
                status_code=status.HTTP_404_NOT_FOUND
            )

        # 4. 檢查假別額度
        current_year = datetime.now().year
        leave_balance = LeaveBalances.objects.filter(
            employee_id=employee,
            year=current_year,
            leave_type=leave_type
        ).first()

        if leave_balance:
            if leave_balance.remaining_hours < D(str(leave_hours)):
                return error_response(
                    f"假別額度不足。剩餘 {leave_balance.remaining_hours} 小時，申請 {leave_hours} 小時",
                    code="INSUFFICIENT_BALANCE",
                    details={
                        'remaining_hours': float(leave_balance.remaining_hours),
                        'requested_hours': float(leave_hours)
                    },
                    status_code=status.HTTP_400_BAD_REQUEST
                )
        else:
            # 如果沒有額度記錄，建立一個預設的
            # 特休 80h, 病假 240h, 事假 112h
            default_hours = {
                'annual': 80.00,
                'sick': 240.00,
                'personal': 112.00,
            }
            total = default_hours.get(leave_type, 0.00)
            leave_balance = LeaveBalances.objects.create(
                employee_id=employee,
                year=current_year,
                leave_type=leave_type,
                total_hours=D(str(total)),
                used_hours=D('0.00'),
                remaining_hours=D(str(total))
            )

        # 5. 建立請假記錄
        leave_data = {
            'relation_id': relation_id,
            'leave_type': leave_type,
            'start_time': start_time,
            'end_time': end_time,
            'leave_hours': leave_hours,
            'leave_reason': leave_reason,
            'status': 'pending'
        }

        if substitute_employee_id:
            leave_data['substitute_employee_id'] = substitute_employee_id

        serializer = LeaveRecordsSerializer(data=leave_data)
        if not serializer.is_valid():
            return validation_error_response(
                "請假資料驗證失敗",
                details=serializer.errors
            )

        leave_record = serializer.save()

        # 6. 根據審批政策自動建立審批記錄
        from .models import ApprovalPolicy, ManagerialRelationship
        import json

        # 計算請假天數
        leave_days = (leave_record.end_time - leave_record.start_time).total_seconds() / (24 * 3600)

        # 查找適用的審批政策
        company_id = relation.company_id
        policy = ApprovalPolicy.objects.filter(
            is_active=True,
            min_days__lte=leave_days
        ).filter(
            models.Q(max_days__gte=leave_days) | models.Q(max_days__isnull=True)
        ).filter(
            models.Q(company_id=company_id) | models.Q(company_id__isnull=True)
        ).order_by('company_id', 'min_days').first()

        if not policy:
            # 預設政策：只需主管審批
            policy_levels = [{"level": 1, "role": "manager", "description": "直屬主管"}]
        else:
            policy_levels = policy.approval_levels if isinstance(policy.approval_levels, list) else json.loads(policy.approval_levels)

        # 根據政策建立審批記錄
        approvals_created = []
        for level_config in policy_levels:
            level = level_config['level']
            role = level_config['role']

            # 根據角色找到審批人
            approver = None
            if role == 'manager':
                # 從直屬主管欄位取得
                approver = relation.direct_manager
                if not approver:
                    # 嘗試從 ManagerialRelationship 取得
                    mgr_rel = ManagerialRelationship.objects.filter(
                        employee_id=employee,
                        effective_date__lte=datetime.now().date(),
                        end_date__isnull=True
                    ).first()
                    if mgr_rel:
                        approver = mgr_rel.manager_id
                    else:
                        # 最後備援：找第一個 MGR 開頭的員工
                        approver = Employees.objects.filter(employee_id__startswith='MGR').first()
            elif role == 'hr':
                approver = Employees.objects.filter(employee_id__startswith='HR').first()
            elif role == 'ceo':
                approver = Employees.objects.filter(employee_id__startswith='CEO').first()

            if approver:
                approval = ApprovalRecords.objects.create(
                    leave_id=leave_record,
                    approver_id=approver,
                    approval_level=level,
                    status='pending'
                )
                approvals_created.append({
                    'id': approval.id,
                    'level': level,
                    'approver': approver.username,
                    'approver_id': approver.employee_id
                })

        if not approvals_created:
            # 如果沒有建立任何審批記錄，使用預設審批人
            default_approver = request.user
            approval = ApprovalRecords.objects.create(
                leave_id=leave_record,
                approver_id=default_approver,
                approval_level=1,
                status='pending'
            )
            approvals_created.append({
                'id': approval.id,
                'level': 1,
                'approver': default_approver.username,
                'approver_id': default_approver.employee_id
            })

        # 7. TODO: 發送通知給審批人（Phase 2 後續實作）

        # 8. 返回成功回應
        return success_response(
            message=f"請假申請已提交，需要 {len(approvals_created)} 層審批",
            data={
                'leave_id': leave_record.id,
                'leave_type': leave_record.get_leave_type_display(),
                'start_time': leave_record.start_time,
                'end_time': leave_record.end_time,
                'leave_hours': float(leave_record.leave_hours),
                'leave_days': round(leave_days, 2),
                'status': leave_record.get_status_display(),
                'approvals': approvals_created,
                'policy_name': policy.policy_name if policy else '預設政策'
            },
            status_code=status.HTTP_201_CREATED
        )

    except Exception as e:
        print(f"請假申請錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response(
            "請假申請失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_leave(request, approval_id):
    """
    批准請假 API（Phase 2 Week 4）

    URL: POST /approval/approve/<approval_id>/
    請求參數：
    - comment: 審批意見（選填）
    """
    try:
        # 1. 取得審批記錄
        try:
            approval = ApprovalRecords.objects.get(id=approval_id)
        except ApprovalRecords.DoesNotExist:
            return not_found_response("審批記錄不存在", code="APPROVAL_NOT_FOUND")

        # 2. 驗證權限：只有審批人可以審批
        if approval.approver_id.employee_id != request.user.employee_id:
            return forbidden_response(
                "您沒有權限審批此申請",
                code="PERMISSION_DENIED"
            )

        # 3. 檢查審批狀態
        if approval.status != 'pending':
            return error_response(
                f"此審批已{approval.get_status_display()}，無法再次審批",
                code="INVALID_STATUS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 更新審批記錄
        approval.status = 'approved'
        approval.comment = request.data.get('comment', '')
        approval.approved_at = timezone.now()
        approval.save()

        # 5. 檢查是否需要下一層級審批
        leave = approval.leave_id
        leave_days = (leave.end_time - leave.start_time).days + 1

        next_level = None
        if leave_days >= 4 and approval.approval_level == 1:
            # 4+ 天，需要 HR 審批
            next_level = 2
        elif leave_days >= 4 and approval.approval_level == 2:
            # 4+ 天，需要總經理審批
            next_level = 3
        elif leave_days >= 2 and approval.approval_level == 1:
            # 2-3 天，需要 HR 審批
            next_level = 2

        if next_level:
            # 建立下一層級審批記錄
            next_approver = Employees.objects.filter(
                employee_id__startswith='HR' if next_level == 2 else 'CEO'
            ).first()

            if next_approver:
                ApprovalRecords.objects.create(
                    leave_id=leave,
                    approver_id=next_approver,
                    approval_level=next_level,
                    status='pending'
                )
                message = f"審批成功，已轉至 Level {next_level} 審批"
            else:
                # 如果找不到下一層級審批人，直接批准請假
                leave.status = 'approved'
                leave.save()
                # 扣除假別額度
                _deduct_leave_balance(leave)
                message = "審批成功，請假已批准"
        else:
            # 最後一層級審批，直接批准請假
            leave.status = 'approved'
            leave.save()

            # 6. 扣除假別額度
            _deduct_leave_balance(leave)

            message = "審批成功，請假已批准"

        # 7. TODO: 發送通知給申請人（Phase 2 後續實作）

        # 8. 返回成功回應
        return success_response(
            message=message,
            data={
                'approval_id': approval.id,
                'leave_id': leave.id,
                'leave_status': leave.get_status_display(),
                'approved_at': approval.approved_at,
                'comment': approval.comment
            }
        )

    except Exception as e:
        print(f"審批錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response(
            "審批失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_leave(request, approval_id):
    """
    拒絕請假 API（Phase 2 Week 4）

    URL: POST /approval/reject/<approval_id>/
    請求參數：
    - comment: 拒絕原因（必填）
    """
    try:
        # 1. 取得審批記錄
        try:
            approval = ApprovalRecords.objects.get(id=approval_id)
        except ApprovalRecords.DoesNotExist:
            return not_found_response("審批記錄不存在", code="APPROVAL_NOT_FOUND")

        # 2. 驗證權限
        if approval.approver_id.employee_id != request.user.employee_id:
            return forbidden_response(
                "您沒有權限審批此申請",
                code="PERMISSION_DENIED"
            )

        # 3. 檢查審批狀態
        if approval.status != 'pending':
            return error_response(
                f"此審批已{approval.get_status_display()}，無法再次審批",
                code="INVALID_STATUS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 驗證拒絕原因
        comment = request.data.get('comment', '').strip()
        if not comment:
            return validation_error_response("請提供拒絕原因")

        # 5. 更新審批記錄
        approval.status = 'rejected'
        approval.comment = comment
        approval.approved_at = timezone.now()
        approval.save()

        # 6. 更新請假記錄狀態
        leave = approval.leave_id
        leave.status = 'rejected'
        leave.save()

        # 7. TODO: 發送通知給申請人（Phase 2 後續實作）

        # 8. 返回成功回應
        return success_response(
            message="已拒絕請假申請",
            data={
                'approval_id': approval.id,
                'leave_id': leave.id,
                'leave_status': leave.get_status_display(),
                'rejected_at': approval.approved_at,
                'comment': approval.comment
            }
        )

    except Exception as e:
        print(f"拒絕審批錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response(
            "拒絕審批失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


def _deduct_leave_balance(leave_record):
    """扣除假別額度（內部函數）"""
    try:
        employee = leave_record.relation_id.employee_id
        year = leave_record.start_time.year

        balance = LeaveBalances.objects.filter(
            employee_id=employee,
            year=year,
            leave_type=leave_record.leave_type
        ).first()

        if balance:
            balance.used_hours += leave_record.leave_hours
            balance.save()  # save() 會自動計算 remaining_hours
            print(f"已扣除 {leave_record.leave_hours} 小時 {leave_record.get_leave_type_display()}")
    except Exception as e:
        print(f"扣除假別額度失敗: {str(e)}")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_leave_records(request):
    """
    查詢我的請假記錄 API（Phase 2 Week 4）

    URL: GET /leave/my-records/
    查詢參數：
    - days: 查詢最近 N 天的記錄（預設 30 天）
    - status: 過濾狀態（pending, approved, rejected, cancelled）
    """
    try:
        user = request.user

        # 取得使用者的所有關聯
        relations = EmpCompanyRel.objects.filter(employee_id=user)

        # 查詢請假記錄
        queryset = LeaveRecords.objects.filter(relation_id__in=relations)

        # 過濾天數
        days = request.query_params.get('days', '30')
        if days:
            days = int(days)
            start_date = datetime.now() - timedelta(days=days)
            queryset = queryset.filter(created_at__gte=start_date)

        # 過濾狀態
        leave_status = request.query_params.get('status')
        if leave_status:
            queryset = queryset.filter(status=leave_status)

        # 排序
        queryset = queryset.order_by('-created_at')

        # 序列化
        serializer = LeaveRecordsSerializer(queryset, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': queryset.count(),
                'records': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢請假記錄錯誤: {str(e)}")
        return server_error_response(
            "查詢失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_balances(request):
    """
    查詢我的假別額度 API（Phase 2 Week 4）

    URL: GET /leave/balances/
    查詢參數：
    - year: 年度（預設當前年度）
    """
    try:
        user = request.user
        year = request.query_params.get('year', datetime.now().year)

        # 查詢假別額度
        balances = LeaveBalances.objects.filter(
            employee_id=user,
            year=year
        ).order_by('leave_type')

        # 序列化
        serializer = LeaveBalancesSerializer(balances, many=True)

        # 計算總額度
        total_hours = sum(b.total_hours for b in balances)
        used_hours = sum(b.used_hours for b in balances)
        remaining_hours = sum(b.remaining_hours for b in balances)

        return success_response(
            message="查詢成功",
            data={
                'year': int(year),
                'summary': {
                    'total_hours': float(total_hours),
                    'used_hours': float(used_hours),
                    'remaining_hours': float(remaining_hours)
                },
                'balances': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢假別額度錯誤: {str(e)}")
        return server_error_response(
            "查詢失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_leave_approvals(request):
    """
    查詢待審批的請假申請（主管用）

    URL: GET /leave/pending/
    """
    try:
        user = request.user

        # 查詢待審批的請假記錄
        approvals = ApprovalRecords.objects.filter(
            approver_id=user,
            status='pending'
        ).select_related('leave_id').order_by('created_at')

        serializer = ApprovalRecordsSerializer(approvals, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': approvals.count(),
                'approvals': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢待審批請假錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_approvals(request):
    """
    查詢待我審批的申請 API（Phase 2 Week 4）

    URL: GET /approval/pending/
    """
    try:
        user = request.user

        # 查詢待審批的記錄
        approvals = ApprovalRecords.objects.filter(
            approver_id=user,
            status='pending'
        ).order_by('created_at')

        # 序列化
        serializer = ApprovalRecordsSerializer(approvals, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': approvals.count(),
                'approvals': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢待審批申請錯誤: {str(e)}")
        return server_error_response(
            "查詢失敗，請稍後再試",
            code="INTERNAL_ERROR"
        )


# =====================================================
# Phase 1 新增：補打卡 API
# =====================================================
from .models import MakeupClockRequest, MakeupClockApproval, MakeupClockQuota, WorkSchedule
from .serializers import MakeupClockRequestSerializer, MakeupClockApprovalSerializer, MakeupClockQuotaSerializer, WorkScheduleSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_makeup_clock(request):
    """
    補打卡申請 API（Phase 1）

    請求參數：
    - relation_id: 員工-公司關聯 ID
    - date: 補打卡日期（YYYY-MM-DD）
    - makeup_type: 補打卡類型（checkin/checkout/both）
    - requested_checkin_time: 申請的上班時間（HH:MM）
    - requested_checkout_time: 申請的下班時間（HH:MM）
    - reason: 補打卡原因
    """
    try:
        # 1. 取得參數
        relation_id = request.data.get('relation_id')
        date_str = request.data.get('date')
        makeup_type = request.data.get('makeup_type', 'checkin')
        requested_checkin = request.data.get('requested_checkin_time')
        requested_checkout = request.data.get('requested_checkout_time')
        reason = request.data.get('reason', '')

        # 2. 參數驗證
        if not all([relation_id, date_str, reason]):
            return validation_error_response("缺少必要參數（relation_id, date, reason）")

        # 3. 驗證 relation
        try:
            relation = EmpCompanyRel.objects.get(id=relation_id)
            employee = relation.employee_id
        except EmpCompanyRel.DoesNotExist:
            return not_found_response("無效的員工-公司關聯")

        # 4. 檢查補打卡額度
        current_year = datetime.now().year
        quota, created = MakeupClockQuota.objects.get_or_create(
            employee_id=employee,
            year=current_year,
            defaults={'total_count': 24, 'used_count': 0}
        )

        if quota.remaining_count <= 0:
            return error_response(
                f"補打卡額度已用完。本年度額度：{quota.total_count} 次",
                code="QUOTA_EXCEEDED",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 5. 檢查日期是否允許（只能補最近 7 天）
        from datetime import date as date_class
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        days_ago = (date_class.today() - target_date).days
        if days_ago > 7 or days_ago < 0:
            return error_response(
                "只能申請最近 7 天內的補打卡",
                code="INVALID_DATE",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 6. 查找原始打卡記錄
        existing_record = AttendanceRecords.objects.filter(
            relation_id=relation,
            date=target_date
        ).first()

        # 7. 處理時間格式
        requested_checkin_dt = None
        requested_checkout_dt = None
        if requested_checkin:
            checkin_time = datetime.strptime(requested_checkin, '%H:%M').time()
            requested_checkin_dt = timezone.make_aware(
                datetime.combine(target_date, checkin_time)
            )
        if requested_checkout:
            checkout_time = datetime.strptime(requested_checkout, '%H:%M').time()
            requested_checkout_dt = timezone.make_aware(
                datetime.combine(target_date, checkout_time)
            )

        # 8. 建立補打卡申請
        makeup_request = MakeupClockRequest.objects.create(
            relation_id=relation,
            date=target_date,
            makeup_type=makeup_type,
            original_checkin_time=existing_record.checkin_time if existing_record else None,
            original_checkout_time=existing_record.checkout_time if existing_record else None,
            requested_checkin_time=requested_checkin_dt,
            requested_checkout_time=requested_checkout_dt,
            reason=reason,
            attendance_record=existing_record
        )

        # 9. 建立審批記錄（直屬主管）
        approver = relation.direct_manager
        if not approver:
            # 嘗試從 ManagerialRelationship 取得
            from .models import ManagerialRelationship
            mgr_rel = ManagerialRelationship.objects.filter(
                employee_id=employee,
                effective_date__lte=date_class.today(),
                end_date__isnull=True
            ).first()
            if mgr_rel:
                approver = mgr_rel.manager_id

        if approver:
            MakeupClockApproval.objects.create(
                request_id=makeup_request,
                approver_id=approver,
                approval_level=1,
                status='pending'
            )

        return success_response(
            message="補打卡申請已提交",
            data={
                'request_id': makeup_request.id,
                'date': str(makeup_request.date),
                'makeup_type': makeup_request.get_makeup_type_display(),
                'status': makeup_request.get_status_display(),
                'remaining_quota': quota.remaining_count
            },
            status_code=status.HTTP_201_CREATED
        )

    except Exception as e:
        print(f"補打卡申請錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("補打卡申請失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_makeup_clock(request, approval_id):
    """
    批准補打卡 API（Phase 1）

    URL: POST /makeup-clock/approve/<approval_id>/
    請求參數：
    - comment: 審批意見（選填）
    """
    try:
        # 1. 取得審批記錄
        try:
            approval = MakeupClockApproval.objects.get(id=approval_id)
        except MakeupClockApproval.DoesNotExist:
            return not_found_response("找不到審批記錄")

        # 2. 驗證權限
        if approval.approver_id != request.user:
            return error_response(
                "您沒有權限審批此申請",
                code="PERMISSION_DENIED",
                status_code=status.HTTP_403_FORBIDDEN
            )

        # 3. 檢查狀態
        if approval.status != 'pending':
            return error_response(
                f"此申請已經{approval.get_status_display()}",
                code="ALREADY_PROCESSED",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 更新審批記錄
        approval.status = 'approved'
        approval.comment = request.data.get('comment', '')
        approval.approved_at = timezone.now()
        approval.save()

        # 5. 更新補打卡申請狀態
        makeup_request = approval.request_id
        makeup_request.status = 'approved'
        makeup_request.save()

        # 6. 建立或更新出勤記錄
        _apply_makeup_clock_to_attendance(makeup_request)

        # 7. 扣除補打卡額度
        employee = makeup_request.relation_id.employee_id
        current_year = datetime.now().year
        try:
            quota = MakeupClockQuota.objects.get(
                employee_id=employee,
                year=current_year
            )
            quota.used_count += 1
            quota.save()
        except MakeupClockQuota.DoesNotExist:
            pass

        return success_response(
            message="已批准補打卡申請",
            data={
                'approval_id': approval.id,
                'request_id': makeup_request.id,
                'status': makeup_request.get_status_display(),
                'approved_at': approval.approved_at
            }
        )

    except Exception as e:
        print(f"批准補打卡錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("批准失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_makeup_clock(request, approval_id):
    """
    拒絕補打卡 API（Phase 1）

    URL: POST /makeup-clock/reject/<approval_id>/
    請求參數：
    - comment: 拒絕原因（必填）
    """
    try:
        # 1. 取得審批記錄
        try:
            approval = MakeupClockApproval.objects.get(id=approval_id)
        except MakeupClockApproval.DoesNotExist:
            return not_found_response("找不到審批記錄")

        # 2. 驗證權限
        if approval.approver_id != request.user:
            return error_response(
                "您沒有權限審批此申請",
                code="PERMISSION_DENIED",
                status_code=status.HTTP_403_FORBIDDEN
            )

        # 3. 檢查狀態
        if approval.status != 'pending':
            return error_response(
                f"此申請已經{approval.get_status_display()}",
                code="ALREADY_PROCESSED",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 驗證拒絕原因
        comment = request.data.get('comment', '')
        if not comment:
            return validation_error_response("請輸入拒絕原因")

        # 5. 更新審批記錄
        approval.status = 'rejected'
        approval.comment = comment
        approval.approved_at = timezone.now()
        approval.save()

        # 6. 更新補打卡申請狀態
        makeup_request = approval.request_id
        makeup_request.status = 'rejected'
        makeup_request.save()

        return success_response(
            message="已拒絕補打卡申請",
            data={
                'approval_id': approval.id,
                'request_id': makeup_request.id,
                'status': makeup_request.get_status_display(),
                'comment': comment
            }
        )

    except Exception as e:
        print(f"拒絕補打卡錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("拒絕失敗，請稍後再試")


def _apply_makeup_clock_to_attendance(makeup_request):
    """將補打卡申請應用到出勤記錄（內部函數）"""
    try:
        relation = makeup_request.relation_id
        target_date = makeup_request.date

        # 查找或建立出勤記錄
        record = makeup_request.attendance_record
        if not record:
            # 建立新的出勤記錄
            record = AttendanceRecords.objects.create(
                relation_id=relation,
                date=target_date,
                checkin_time=makeup_request.requested_checkin_time or timezone.now(),
                checkout_time=makeup_request.requested_checkout_time or timezone.now(),
                checkin_location="補打卡",
                checkout_location="補打卡",
                work_hours=Decimal('0.00'),
                is_makeup=True
            )

        # 根據補打卡類型更新記錄
        if makeup_request.makeup_type in ['checkin', 'both']:
            if makeup_request.requested_checkin_time:
                record.checkin_time = makeup_request.requested_checkin_time
                record.checkin_location = "補打卡"

        if makeup_request.makeup_type in ['checkout', 'both']:
            if makeup_request.requested_checkout_time:
                record.checkout_time = makeup_request.requested_checkout_time
                record.checkout_location = "補打卡"

        # 重新計算工時
        if record.checkin_time and record.checkout_time:
            record.work_hours = calculate_work_hours(record.checkin_time, record.checkout_time)

        record.is_makeup = True
        record.save()

        # 更新 makeup_request 的關聯
        makeup_request.attendance_record = record
        makeup_request.save()

        print(f"補打卡已應用到出勤記錄 #{record.id}")

    except Exception as e:
        print(f"應用補打卡失敗: {str(e)}")
        import traceback
        traceback.print_exc()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_makeup_requests(request):
    """
    查詢我的補打卡申請 API（Phase 1）

    URL: GET /makeup-clock/my-requests/
    查詢參數：
    - days: 查詢最近 N 天的記錄（預設 30 天）
    - status: 過濾狀態（pending, approved, rejected）
    """
    try:
        user = request.user

        # 取得使用者的所有關聯
        relations = EmpCompanyRel.objects.filter(employee_id=user)

        # 查詢補打卡記錄
        queryset = MakeupClockRequest.objects.filter(relation_id__in=relations)

        # 過濾天數
        days = request.query_params.get('days', '30')
        if days:
            days = int(days)
            start_date = datetime.now() - timedelta(days=days)
            queryset = queryset.filter(created_at__gte=start_date)

        # 過濾狀態
        req_status = request.query_params.get('status')
        if req_status:
            queryset = queryset.filter(status=req_status)

        # 排序
        queryset = queryset.order_by('-created_at')

        # 序列化
        serializer = MakeupClockRequestSerializer(queryset, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': queryset.count(),
                'requests': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢補打卡記錄錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def makeup_clock_quota(request):
    """
    查詢我的補打卡額度 API（Phase 1）

    URL: GET /makeup-clock/quota/
    查詢參數：
    - year: 年度（預設當前年度）
    """
    try:
        user = request.user
        year = request.query_params.get('year', datetime.now().year)

        # 查詢或建立額度
        quota, created = MakeupClockQuota.objects.get_or_create(
            employee_id=user,
            year=int(year),
            defaults={'total_count': 24, 'used_count': 0}
        )

        return success_response(
            message="查詢成功",
            data={
                'year': int(year),
                'total_count': quota.total_count,
                'used_count': quota.used_count,
                'remaining_count': quota.remaining_count
            }
        )

    except Exception as e:
        print(f"查詢補打卡額度錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_makeup_approvals(request):
    """
    查詢待我審批的補打卡申請 API（Phase 1）

    URL: GET /makeup-clock/pending/
    """
    try:
        user = request.user

        # 查詢待審批的記錄
        approvals = MakeupClockApproval.objects.filter(
            approver_id=user,
            status='pending'
        ).order_by('created_at')

        # 序列化
        serializer = MakeupClockApprovalSerializer(approvals, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': approvals.count(),
                'approvals': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢待審批補打卡錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_work_schedule(request):
    """
    查詢我的班表 API（Phase 1）

    URL: GET /schedule/my-schedule/
    """
    try:
        user = request.user

        # 取得使用者的關聯
        relation = EmpCompanyRel.objects.filter(
            employee_id=user,
            employment_status=True
        ).first()

        if not relation:
            return not_found_response("找不到員工關聯")

        # 取得班表（優先員工專屬，否則使用公司預設）
        schedule = relation.work_schedule
        if not schedule:
            schedule = WorkSchedule.objects.filter(
                company_id=relation.company_id,
                is_default=True,
                is_active=True
            ).first()

        if not schedule:
            return success_response(
                message="尚未設定班表",
                data=None
            )

        serializer = WorkScheduleSerializer(schedule)

        return success_response(
            message="查詢成功",
            data=serializer.data
        )

    except Exception as e:
        print(f"查詢班表錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


# =====================================================
# Phase 2 新增：加班管理 API
# =====================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_overtime(request):
    """
    申請加班 API

    URL: POST /overtime/apply/
    請求參數：
    - date: 加班日期 (YYYY-MM-DD)
    - start_time: 開始時間 (HH:MM)
    - end_time: 結束時間 (HH:MM)
    - reason: 加班原因
    - compensation_type: 補償方式 (pay/compensatory/mixed)
    - compensatory_hours: 補休時數（選填，compensation_type 為 compensatory 或 mixed 時）
    - pay_hours: 加班費時數（選填，compensation_type 為 pay 或 mixed 時）
    """
    try:
        user = request.user
        data = request.data

        # 1. 驗證必填欄位
        required_fields = ['date', 'start_time', 'end_time', 'reason']
        for field in required_fields:
            if not data.get(field):
                return validation_error_response(f"缺少必填欄位：{field}")

        # 2. 取得員工關聯
        relation = EmpCompanyRel.objects.filter(
            employee_id=user,
            employment_status=True
        ).first()

        if not relation:
            return not_found_response("找不到員工關聯")

        # 3. 解析日期和時間
        try:
            ot_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
            start_time = datetime.strptime(data['start_time'], '%H:%M').time()
            end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        except ValueError:
            return validation_error_response("日期或時間格式錯誤")

        # 4. 計算加班時數
        start_dt = datetime.combine(ot_date, start_time)
        end_dt = datetime.combine(ot_date, end_time)
        if end_dt <= start_dt:
            # 跨日加班
            end_dt = end_dt + timedelta(days=1)
        overtime_hours = Decimal(str((end_dt - start_dt).total_seconds() / 3600))

        # 5. 取得補償方式
        compensation_type = data.get('compensation_type', 'compensatory')
        if compensation_type not in ['pay', 'compensatory', 'mixed']:
            compensation_type = 'compensatory'

        # 計算補休和加班費時數
        if compensation_type == 'compensatory':
            compensatory_hours = overtime_hours
            pay_hours = Decimal('0')
        elif compensation_type == 'pay':
            compensatory_hours = Decimal('0')
            pay_hours = overtime_hours
        else:  # mixed
            compensatory_hours = Decimal(str(data.get('compensatory_hours', 0)))
            pay_hours = Decimal(str(data.get('pay_hours', 0)))
            if compensatory_hours + pay_hours != overtime_hours:
                compensatory_hours = overtime_hours
                pay_hours = Decimal('0')

        # 6. 建立加班記錄
        overtime_record = OvertimeRecords.objects.create(
            relation_id=relation,
            date=ot_date,
            start_time=start_time,
            end_time=end_time,
            overtime_hours=overtime_hours,
            reason=data['reason'],
            compensation_type=compensation_type,
            compensatory_hours=compensatory_hours,
            pay_hours=pay_hours,
            status='pending'
        )

        # 7. 建立審批記錄（直屬主管）
        approver = relation.direct_manager
        if not approver:
            approver = Employees.objects.filter(employee_id__startswith='MGR').first()
        if not approver:
            approver = request.user

        approval = OvertimeApproval.objects.create(
            overtime_id=overtime_record,
            approver_id=approver,
            approval_level=1,
            status='pending'
        )

        # 8. 建立通知
        Notifications.objects.create(
            recipient_id=approver,
            notification_type='approval_pending',
            title='新加班申請待審批',
            content=f'{user.username} 申請 {ot_date} 加班 {overtime_hours} 小時',
            related_model='OvertimeRecords',
            related_id=overtime_record.id
        )

        return success_response(
            message="加班申請已提交",
            data={
                'id': overtime_record.id,
                'date': str(overtime_record.date),
                'start_time': str(overtime_record.start_time),
                'end_time': str(overtime_record.end_time),
                'overtime_hours': float(overtime_record.overtime_hours),
                'compensation_type': overtime_record.get_compensation_type_display(),
                'status': overtime_record.get_status_display(),
                'approval': {
                    'id': approval.id,
                    'approver': approver.username
                }
            },
            status_code=status.HTTP_201_CREATED
        )

    except Exception as e:
        print(f"加班申請錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("加班申請失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_overtime_records(request):
    """
    查詢我的加班記錄 API

    URL: GET /overtime/my-records/
    查詢參數：
    - status: 狀態篩選（選填）
    - days: 查詢天數（選填，預設 30）
    """
    try:
        user = request.user
        status_filter = request.query_params.get('status')
        days = int(request.query_params.get('days', 30))

        # 取得員工關聯
        relations = EmpCompanyRel.objects.filter(
            employee_id=user
        ).values_list('id', flat=True)

        # 查詢加班記錄
        queryset = OvertimeRecords.objects.filter(
            relation_id__in=relations
        )

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # 時間篩選
        start_date = timezone.now().date() - timedelta(days=days)
        queryset = queryset.filter(date__gte=start_date)

        serializer = OvertimeRecordsSerializer(queryset, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': queryset.count(),
                'records': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢加班記錄錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_overtime(request, approval_id):
    """
    批准加班申請 API

    URL: POST /overtime/approve/<approval_id>/
    請求參數：
    - comment: 審批意見（選填）
    """
    try:
        # 1. 取得審批記錄
        try:
            approval = OvertimeApproval.objects.get(id=approval_id)
        except OvertimeApproval.DoesNotExist:
            return not_found_response("審批記錄不存在")

        # 2. 驗證權限
        if approval.approver_id.employee_id != request.user.employee_id:
            return forbidden_response("您沒有權限審批此申請")

        # 3. 檢查審批狀態
        if approval.status != 'pending':
            return error_response(
                f"此審批已{approval.get_status_display()}",
                code="INVALID_STATUS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 更新審批記錄
        approval.status = 'approved'
        approval.comment = request.data.get('comment', '')
        approval.approved_at = timezone.now()
        approval.save()

        # 5. 更新加班記錄狀態
        overtime = approval.overtime_id
        overtime.status = 'approved'
        overtime.save()

        # 6. 更新補休額度（如果選擇補休）
        if overtime.compensatory_hours > 0:
            employee = overtime.relation_id.employee_id
            year = overtime.date.year

            # 取得或建立補休額度
            balance, created = LeaveBalances.objects.get_or_create(
                employee_id=employee,
                year=year,
                leave_type='compensatory',
                defaults={
                    'total_hours': Decimal('0'),
                    'used_hours': Decimal('0'),
                    'remaining_hours': Decimal('0')
                }
            )
            balance.total_hours += overtime.compensatory_hours
            balance.remaining_hours = balance.total_hours - balance.used_hours
            balance.save()

        # 7. 建立通知
        applicant = overtime.relation_id.employee_id
        Notifications.objects.create(
            recipient_id=applicant,
            notification_type='approval_result',
            title='加班申請已批准',
            content=f'您 {overtime.date} 的加班申請已獲批准',
            related_model='OvertimeRecords',
            related_id=overtime.id
        )

        return success_response(
            message="已批准加班申請",
            data={
                'overtime_id': overtime.id,
                'status': overtime.get_status_display(),
                'compensatory_hours_added': float(overtime.compensatory_hours)
            }
        )

    except Exception as e:
        print(f"批准加班錯誤: {str(e)}")
        return server_error_response("操作失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_overtime(request, approval_id):
    """
    拒絕加班申請 API

    URL: POST /overtime/reject/<approval_id>/
    請求參數：
    - comment: 拒絕原因（必填）
    """
    try:
        comment = request.data.get('comment')
        if not comment:
            return validation_error_response("請填寫拒絕原因")

        # 1. 取得審批記錄
        try:
            approval = OvertimeApproval.objects.get(id=approval_id)
        except OvertimeApproval.DoesNotExist:
            return not_found_response("審批記錄不存在")

        # 2. 驗證權限
        if approval.approver_id.employee_id != request.user.employee_id:
            return forbidden_response("您沒有權限審批此申請")

        # 3. 檢查審批狀態
        if approval.status != 'pending':
            return error_response(
                f"此審批已{approval.get_status_display()}",
                code="INVALID_STATUS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 更新審批記錄
        approval.status = 'rejected'
        approval.comment = comment
        approval.approved_at = timezone.now()
        approval.save()

        # 5. 更新加班記錄狀態
        overtime = approval.overtime_id
        overtime.status = 'rejected'
        overtime.save()

        # 6. 建立通知
        applicant = overtime.relation_id.employee_id
        Notifications.objects.create(
            recipient_id=applicant,
            notification_type='approval_result',
            title='加班申請已拒絕',
            content=f'您 {overtime.date} 的加班申請已被拒絕。原因：{comment}',
            related_model='OvertimeRecords',
            related_id=overtime.id
        )

        return success_response(
            message="已拒絕加班申請",
            data={
                'overtime_id': overtime.id,
                'status': overtime.get_status_display()
            }
        )

    except Exception as e:
        print(f"拒絕加班錯誤: {str(e)}")
        return server_error_response("操作失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_overtime_approvals(request):
    """
    查詢待審批的加班申請（主管用）

    URL: GET /overtime/pending/
    """
    try:
        user = request.user

        # 查詢待審批記錄
        approvals = OvertimeApproval.objects.filter(
            approver_id=user,
            status='pending'
        ).select_related('overtime_id')

        serializer = OvertimeApprovalSerializer(approvals, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': approvals.count(),
                'approvals': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢待審批加班錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cancel_overtime(request, overtime_id):
    """
    取消加班申請 API

    URL: POST /overtime/cancel/<overtime_id>/
    """
    try:
        user = request.user

        # 1. 取得加班記錄
        try:
            overtime = OvertimeRecords.objects.get(id=overtime_id)
        except OvertimeRecords.DoesNotExist:
            return not_found_response("加班記錄不存在")

        # 2. 驗證權限（只能取消自己的申請）
        if overtime.relation_id.employee_id.employee_id != user.employee_id:
            return forbidden_response("您只能取消自己的申請")

        # 3. 檢查狀態（只能取消待審批的申請）
        if overtime.status != 'pending':
            return error_response(
                f"此申請已{overtime.get_status_display()}，無法取消",
                code="INVALID_STATUS",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 4. 更新狀態
        overtime.status = 'cancelled'
        overtime.save()

        # 更新審批記錄
        OvertimeApproval.objects.filter(
            overtime_id=overtime,
            status='pending'
        ).update(status='rejected', comment='申請人已取消')

        return success_response(
            message="已取消加班申請",
            data={'overtime_id': overtime.id}
        )

    except Exception as e:
        print(f"取消加班錯誤: {str(e)}")
        return server_error_response("操作失敗，請稍後再試")


# =====================================================
# Phase 2 新增：特休自動計算 API
# =====================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def annual_leave_entitlement(request):
    """
    查詢特休資格明細 API

    URL: GET /leave/annual-entitlement/
    """
    try:
        from .utils import calculate_annual_leave_days

        user = request.user

        # 取得員工關聯（取得入職日期）
        relation = EmpCompanyRel.objects.filter(
            employee_id=user,
            employment_status=True
        ).first()

        if not relation:
            return not_found_response("找不到員工關聯")

        if not relation.hire_date:
            return error_response(
                "尚未設定入職日期",
                code="NO_HIRE_DATE",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 計算特休
        result = calculate_annual_leave_days(relation.hire_date)

        # 取得今年已使用的特休
        year = timezone.now().year
        balance = LeaveBalances.objects.filter(
            employee_id=user,
            year=year,
            leave_type='annual'
        ).first()

        used_hours = float(balance.used_hours) if balance else 0
        total_hours = result['hours']
        remaining_hours = total_hours - used_hours

        return success_response(
            message="查詢成功",
            data={
                'hire_date': str(relation.hire_date),
                'seniority': {
                    'years': result['years'],
                    'months': result['months']
                },
                'annual_leave': {
                    'days': result['days'],
                    'hours': total_hours,
                    'used_hours': used_hours,
                    'remaining_hours': max(0, remaining_hours),
                    'description': result['description']
                }
            }
        )

    except Exception as e:
        print(f"查詢特休資格錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculate_annual_leave(request):
    """
    計算並更新特休額度 API

    URL: POST /leave/calculate-annual/
    請求參數：
    - year: 年度（選填，預設今年）
    """
    try:
        from .utils import calculate_annual_leave_days

        user = request.user
        year = int(request.data.get('year', timezone.now().year))

        # 取得員工關聯
        relation = EmpCompanyRel.objects.filter(
            employee_id=user,
            employment_status=True
        ).first()

        if not relation:
            return not_found_response("找不到員工關聯")

        if not relation.hire_date:
            return error_response(
                "尚未設定入職日期",
                code="NO_HIRE_DATE",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 計算特休
        result = calculate_annual_leave_days(relation.hire_date)
        total_hours = Decimal(str(result['hours']))

        # 更新或建立假別額度
        balance, created = LeaveBalances.objects.get_or_create(
            employee_id=user,
            year=year,
            leave_type='annual',
            defaults={
                'total_hours': total_hours,
                'used_hours': Decimal('0'),
                'remaining_hours': total_hours
            }
        )

        if not created:
            # 更新總額度（保留已使用時數）
            balance.total_hours = total_hours
            balance.remaining_hours = total_hours - balance.used_hours
            balance.save()

        return success_response(
            message="特休額度已更新" if not created else "特休額度已建立",
            data={
                'year': year,
                'total_hours': float(balance.total_hours),
                'used_hours': float(balance.used_hours),
                'remaining_hours': float(balance.remaining_hours),
                'calculation': {
                    'seniority_years': result['years'],
                    'seniority_months': result['months'],
                    'annual_leave_days': result['days'],
                    'description': result['description']
                }
            }
        )

    except Exception as e:
        print(f"計算特休錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("計算失敗，請稍後再試")


# =====================================================
# Phase 2 新增：出勤報表 API
# =====================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attendance_summary(request):
    """
    個人出勤摘要 API

    URL: GET /reports/attendance-summary/
    查詢參數：
    - year: 年度（選填，預設今年）
    - month: 月份（選填，預設當月）
    """
    try:
        user = request.user
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        # 取得員工關聯
        relations = EmpCompanyRel.objects.filter(
            employee_id=user
        ).values_list('id', flat=True)

        # 計算日期範圍
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, last_day)

        # 查詢出勤記錄
        records = AttendanceRecords.objects.filter(
            relation_id__in=relations,
            date__gte=start_date,
            date__lte=end_date
        )

        # 統計
        total_records = records.count()
        late_records = records.filter(is_late=True)
        early_leave_records = records.filter(is_early_leave=True)
        makeup_records = records.filter(is_makeup=True)

        late_count = late_records.count()
        late_minutes_total = sum(r.late_minutes for r in late_records)
        early_leave_count = early_leave_records.count()
        early_leave_minutes_total = sum(r.early_leave_minutes for r in early_leave_records)

        # 計算總工時
        total_work_hours = sum(
            float(r.work_hours or 0) for r in records
        )

        # 查詢請假時數
        leave_hours = LeaveRecords.objects.filter(
            relation_id__in=relations,
            start_time__year=year,
            start_time__month=month,
            status='approved'
        ).aggregate(total=models.Sum('leave_hours'))['total'] or 0

        # 查詢加班時數
        overtime_hours = OvertimeRecords.objects.filter(
            relation_id__in=relations,
            date__year=year,
            date__month=month,
            status='approved'
        ).aggregate(total=models.Sum('overtime_hours'))['total'] or 0

        return success_response(
            message="查詢成功",
            data={
                'period': {
                    'year': year,
                    'month': month,
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'attendance': {
                    'total_days': total_records,
                    'late_count': late_count,
                    'late_minutes_total': late_minutes_total,
                    'early_leave_count': early_leave_count,
                    'early_leave_minutes_total': early_leave_minutes_total,
                    'makeup_count': makeup_records.count(),
                    'total_work_hours': round(total_work_hours, 2)
                },
                'leave': {
                    'total_hours': float(leave_hours)
                },
                'overtime': {
                    'total_hours': float(overtime_hours)
                }
            }
        )

    except Exception as e:
        print(f"查詢出勤摘要錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def anomaly_list(request):
    """
    出勤異常清單 API

    URL: GET /reports/anomaly-list/
    查詢參數：
    - year: 年度（選填）
    - month: 月份（選填）
    - type: 異常類型（late/early_leave/all，預設 all）
    """
    try:
        user = request.user
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        anomaly_type = request.query_params.get('type', 'all')

        # 取得員工關聯
        relations = EmpCompanyRel.objects.filter(
            employee_id=user
        ).values_list('id', flat=True)

        # 查詢異常記錄
        queryset = AttendanceRecords.objects.filter(
            relation_id__in=relations
        )

        if anomaly_type == 'late':
            queryset = queryset.filter(is_late=True)
        elif anomaly_type == 'early_leave':
            queryset = queryset.filter(is_early_leave=True)
        else:
            queryset = queryset.filter(
                models.Q(is_late=True) | models.Q(is_early_leave=True)
            )

        if year:
            queryset = queryset.filter(date__year=int(year))
        if month:
            queryset = queryset.filter(date__month=int(month))

        # 排序
        queryset = queryset.order_by('-date')[:50]

        # 整理結果
        anomalies = []
        for record in queryset:
            anomaly_info = {
                'id': record.id,
                'date': str(record.date),
                'checkin_time': str(record.checkin_time) if record.checkin_time else None,
                'checkout_time': str(record.checkout_time) if record.checkout_time else None,
                'anomalies': []
            }
            if record.is_late:
                anomaly_info['anomalies'].append({
                    'type': 'late',
                    'description': f'遲到 {record.late_minutes} 分鐘'
                })
            if record.is_early_leave:
                anomaly_info['anomalies'].append({
                    'type': 'early_leave',
                    'description': f'早退 {record.early_leave_minutes} 分鐘'
                })
            anomalies.append(anomaly_info)

        return success_response(
            message="查詢成功",
            data={
                'count': len(anomalies),
                'anomalies': anomalies
            }
        )

    except Exception as e:
        print(f"查詢異常清單錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


# =====================================================
# Phase 2 新增：通知系統 API
# =====================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_notifications(request):
    """
    取得通知列表 API

    URL: GET /notifications/
    查詢參數：
    - unread_only: 只顯示未讀（true/false，預設 false）
    - limit: 數量限制（預設 20）
    """
    try:
        user = request.user
        unread_only = request.query_params.get('unread_only', 'false').lower() == 'true'
        limit = int(request.query_params.get('limit', 20))

        queryset = Notifications.objects.filter(recipient_id=user)

        if unread_only:
            queryset = queryset.filter(is_read=False)

        queryset = queryset.order_by('-created_at')[:limit]

        serializer = NotificationsSerializer(queryset, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': queryset.count(),
                'notifications': serializer.data
            }
        )

    except Exception as e:
        print(f"查詢通知錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def unread_notification_count(request):
    """
    取得未讀通知數量 API

    URL: GET /notifications/unread-count/
    """
    try:
        user = request.user
        count = Notifications.objects.filter(
            recipient_id=user,
            is_read=False
        ).count()

        return success_response(
            message="查詢成功",
            data={'unread_count': count}
        )

    except Exception as e:
        print(f"查詢未讀數量錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, notification_id):
    """
    標記通知為已讀 API

    URL: POST /notifications/mark-read/<notification_id>/
    """
    try:
        user = request.user

        try:
            notification = Notifications.objects.get(
                id=notification_id,
                recipient_id=user
            )
        except Notifications.DoesNotExist:
            return not_found_response("通知不存在")

        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()

        return success_response(
            message="已標記為已讀",
            data={'notification_id': notification.id}
        )

    except Exception as e:
        print(f"標記已讀錯誤: {str(e)}")
        return server_error_response("操作失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_all_notifications_read(request):
    """
    標記所有通知為已讀 API

    URL: POST /notifications/mark-all-read/
    """
    try:
        user = request.user

        updated = Notifications.objects.filter(
            recipient_id=user,
            is_read=False
        ).update(
            is_read=True,
            read_at=timezone.now()
        )

        return success_response(
            message=f"已標記 {updated} 則通知為已讀",
            data={'updated_count': updated}
        )

    except Exception as e:
        print(f"標記全部已讀錯誤: {str(e)}")
        return server_error_response("操作失敗，請稍後再試")


# =====================================================
# Phase 3 新增：使用者資訊與角色權限 API
# =====================================================

from .serializers import UserProfileSerializer, DepartmentsSerializer, EmployeeListSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_profile(request):
    """
    取得當前使用者完整資訊（含角色、權限）

    URL: GET /api/user/profile/
    """
    try:
        user = request.user
        serializer = UserProfileSerializer(user)

        return success_response(
            message="查詢成功",
            data=serializer.data
        )

    except Exception as e:
        print(f"取得使用者資訊錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


# =====================================================
# Phase 3 新增：主管儀表板 API
# =====================================================

def _check_manager_permission(user):
    """檢查是否有主管權限"""
    return user.role in ['manager', 'hr_admin', 'ceo', 'system_admin']


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def manager_dashboard(request):
    """
    主管儀表板 - 部門出勤總覽

    URL: GET /api/manager/dashboard/
    查詢參數：
    - date: 查詢日期（預設今天）
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_manager_permission(user):
            return forbidden_response("您沒有權限存取此功能")

        # 取得查詢日期
        date_str = request.query_params.get('date')
        if date_str:
            query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            query_date = date.today()

        # 取得下屬員工
        subordinates = _get_subordinates(user)
        subordinate_ids = [e.employee_id for e in subordinates]

        # 取得下屬的關聯 ID
        relations = EmpCompanyRel.objects.filter(
            employee_id__in=subordinate_ids,
            employment_status=True
        )
        relation_ids = relations.values_list('id', flat=True)

        # 統計出勤資訊
        attendance_records = AttendanceRecords.objects.filter(
            relation_id__in=relation_ids,
            date=query_date
        )

        # 統計數據
        total_employees = len(subordinates)
        checked_in = attendance_records.count()
        late_count = attendance_records.filter(is_late=True).count()
        early_leave_count = attendance_records.filter(is_early_leave=True).count()

        # 取得待審批數量
        pending_leave = ApprovalRecords.objects.filter(
            approver_id=user,
            status='pending'
        ).count()

        pending_overtime = OvertimeApproval.objects.filter(
            approver_id=user,
            status='pending'
        ).count()

        pending_makeup = MakeupClockApproval.objects.filter(
            approver_id=user,
            status='pending'
        ).count()

        # 取得未打卡員工
        checked_in_employees = set(
            AttendanceRecords.objects.filter(
                relation_id__in=relation_ids,
                date=query_date
            ).values_list('relation_id__employee_id', flat=True)
        )
        not_checked_in = [
            {'employee_id': e.employee_id, 'username': e.username}
            for e in subordinates
            if e.employee_id not in checked_in_employees
        ]

        return success_response(
            message="查詢成功",
            data={
                'date': str(query_date),
                'summary': {
                    'total_employees': total_employees,
                    'checked_in': checked_in,
                    'not_checked_in': total_employees - checked_in,
                    'late_count': late_count,
                    'early_leave_count': early_leave_count,
                },
                'pending_approvals': {
                    'leave': pending_leave,
                    'overtime': pending_overtime,
                    'makeup': pending_makeup,
                    'total': pending_leave + pending_overtime + pending_makeup,
                },
                'not_checked_in_list': not_checked_in[:10],  # 最多顯示 10 人
            }
        )

    except Exception as e:
        print(f"主管儀表板錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("查詢失敗，請稍後再試")


def _get_subordinates(manager):
    """取得主管的下屬員工"""
    from .models import Departments

    subordinates = set()

    # 1. 從 EmpCompanyRel 的 direct_manager 取得直屬下屬
    direct_reports = EmpCompanyRel.objects.filter(
        direct_manager=manager,
        employment_status=True
    ).select_related('employee_id')
    for rel in direct_reports:
        subordinates.add(rel.employee_id)

    # 2. 如果是部門主管，取得部門內所有員工
    if manager.department:
        dept_employees = Employees.objects.filter(
            department=manager.department,
            is_active=True
        ).exclude(employee_id=manager.employee_id)
        subordinates.update(dept_employees)

    # 3. 如果是 HR 或 CEO，取得所有員工
    if manager.role in ['hr_admin', 'ceo', 'system_admin']:
        all_employees = Employees.objects.filter(is_active=True)
        subordinates.update(all_employees)

    return list(subordinates)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def department_report(request):
    """
    部門員工出勤統計

    URL: GET /api/manager/reports/department/
    查詢參數：
    - year: 年份
    - month: 月份
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_manager_permission(user):
            return forbidden_response("您沒有權限存取此功能")

        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        # 計算日期範圍
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, last_day)

        # 取得下屬
        subordinates = _get_subordinates(user)
        subordinate_ids = [e.employee_id for e in subordinates]

        # 取得關聯
        relations = EmpCompanyRel.objects.filter(
            employee_id__in=subordinate_ids,
            employment_status=True
        )

        # 統計每個員工的出勤
        report_data = []
        for rel in relations:
            emp = rel.employee_id
            records = AttendanceRecords.objects.filter(
                relation_id=rel,
                date__gte=start_date,
                date__lte=end_date
            )

            total_days = records.count()
            late_count = records.filter(is_late=True).count()
            early_leave_count = records.filter(is_early_leave=True).count()
            total_work_hours = sum(float(r.work_hours or 0) for r in records)
            late_minutes_total = sum(r.late_minutes for r in records.filter(is_late=True))

            # 請假統計
            leave_hours = LeaveRecords.objects.filter(
                relation_id=rel,
                start_time__year=year,
                start_time__month=month,
                status='approved'
            ).aggregate(total=models.Sum('leave_hours'))['total'] or 0

            # 加班統計
            overtime_hours = OvertimeRecords.objects.filter(
                relation_id=rel,
                date__year=year,
                date__month=month,
                status='approved'
            ).aggregate(total=models.Sum('overtime_hours'))['total'] or 0

            report_data.append({
                'employee_id': emp.employee_id,
                'username': emp.username,
                'department': emp.department.name if emp.department else None,
                'attendance': {
                    'total_days': total_days,
                    'late_count': late_count,
                    'late_minutes_total': late_minutes_total,
                    'early_leave_count': early_leave_count,
                    'total_work_hours': round(total_work_hours, 2),
                },
                'leave_hours': float(leave_hours),
                'overtime_hours': float(overtime_hours),
            })

        return success_response(
            message="查詢成功",
            data={
                'period': {
                    'year': year,
                    'month': month,
                    'start_date': str(start_date),
                    'end_date': str(end_date),
                },
                'employee_count': len(report_data),
                'employees': report_data,
            }
        )

    except Exception as e:
        print(f"部門報表錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def batch_approve(request):
    """
    批次審批 API

    URL: POST /api/approval/batch/
    請求參數：
    - approval_type: 審批類型（leave/overtime/makeup）
    - approval_ids: 審批記錄 ID 陣列
    - action: 動作（approve/reject）
    - comment: 審批意見（選填，拒絕時必填）
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_manager_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        approval_type = request.data.get('approval_type')
        approval_ids = request.data.get('approval_ids', [])
        action = request.data.get('action')
        comment = request.data.get('comment', '')

        if not approval_type or not approval_ids or not action:
            return validation_error_response("缺少必要參數")

        if action not in ['approve', 'reject']:
            return validation_error_response("無效的操作類型")

        if action == 'reject' and not comment:
            return validation_error_response("拒絕時必須填寫原因")

        # 根據類型取得對應的審批模型
        model_map = {
            'leave': ApprovalRecords,
            'overtime': OvertimeApproval,
            'makeup': MakeupClockApproval,
        }

        if approval_type not in model_map:
            return validation_error_response("無效的審批類型")

        ApprovalModel = model_map[approval_type]

        # 批次處理
        processed = []
        failed = []

        for approval_id in approval_ids:
            try:
                approval = ApprovalModel.objects.get(id=approval_id)

                # 檢查權限
                if approval.approver_id != user:
                    failed.append({
                        'id': approval_id,
                        'reason': '無權限審批'
                    })
                    continue

                # 檢查狀態
                if approval.status != 'pending':
                    failed.append({
                        'id': approval_id,
                        'reason': '已處理過'
                    })
                    continue

                # 更新審批記錄
                approval.status = 'approved' if action == 'approve' else 'rejected'
                approval.comment = comment
                approval.approved_at = timezone.now()
                approval.save()

                # 更新關聯記錄
                if approval_type == 'leave':
                    leave = approval.leave_id
                    if action == 'approve':
                        leave.status = 'approved'
                        _deduct_leave_balance(leave)
                    else:
                        leave.status = 'rejected'
                    leave.save()
                elif approval_type == 'overtime':
                    overtime = approval.overtime_id
                    if action == 'approve':
                        overtime.status = 'approved'
                        # 更新補休額度
                        if overtime.compensatory_hours > 0:
                            employee = overtime.relation_id.employee_id
                            year = overtime.date.year
                            balance, _ = LeaveBalances.objects.get_or_create(
                                employee_id=employee,
                                year=year,
                                leave_type='compensatory',
                                defaults={
                                    'total_hours': Decimal('0'),
                                    'used_hours': Decimal('0'),
                                    'remaining_hours': Decimal('0')
                                }
                            )
                            balance.total_hours += overtime.compensatory_hours
                            balance.remaining_hours = balance.total_hours - balance.used_hours
                            balance.save()
                    else:
                        overtime.status = 'rejected'
                    overtime.save()
                elif approval_type == 'makeup':
                    makeup_request = approval.request_id
                    if action == 'approve':
                        makeup_request.status = 'approved'
                        _apply_makeup_clock_to_attendance(makeup_request)
                        # 扣除額度
                        employee = makeup_request.relation_id.employee_id
                        current_year = datetime.now().year
                        try:
                            quota = MakeupClockQuota.objects.get(
                                employee_id=employee,
                                year=current_year
                            )
                            quota.used_count += 1
                            quota.save()
                        except MakeupClockQuota.DoesNotExist:
                            pass
                    else:
                        makeup_request.status = 'rejected'
                    makeup_request.save()

                processed.append(approval_id)

            except ApprovalModel.DoesNotExist:
                failed.append({
                    'id': approval_id,
                    'reason': '記錄不存在'
                })

        return success_response(
            message=f"批次審批完成：成功 {len(processed)} 筆，失敗 {len(failed)} 筆",
            data={
                'processed': processed,
                'failed': failed,
            }
        )

    except Exception as e:
        print(f"批次審批錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("批次審批失敗，請稍後再試")


# =====================================================
# Phase 3 新增：HR 管理 API
# =====================================================

def _check_hr_permission(user):
    """檢查是否有 HR 權限"""
    return user.role in ['hr_admin', 'ceo', 'system_admin']


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hr_employee_list(request):
    """
    HR 員工列表（分頁、篩選）

    URL: GET /api/hr/employees/
    查詢參數：
    - page: 頁碼
    - page_size: 每頁筆數（預設 20）
    - search: 搜尋（員工編號、姓名）
    - department: 部門 ID
    - role: 角色
    - is_active: 是否在職
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限存取此功能")

        # 取得參數
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        search = request.query_params.get('search', '')
        department_id = request.query_params.get('department')
        role = request.query_params.get('role')
        is_active = request.query_params.get('is_active')

        # 建立查詢
        queryset = Employees.objects.all()

        # 搜尋
        if search:
            queryset = queryset.filter(
                Q(employee_id__icontains=search) |
                Q(username__icontains=search) |
                Q(email__icontains=search)
            )

        # 部門篩選
        if department_id:
            queryset = queryset.filter(department_id=department_id)

        # 角色篩選
        if role:
            queryset = queryset.filter(role=role)

        # 在職篩選
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)

        # 排序
        queryset = queryset.order_by('employee_id')

        # 計算總數
        total = queryset.count()

        # 分頁
        start = (page - 1) * page_size
        end = start + page_size
        employees = queryset[start:end]

        # 序列化
        serializer = EmployeeListSerializer(employees, many=True)

        return success_response(
            message="查詢成功",
            data={
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size,
                'employees': serializer.data,
            }
        )

    except Exception as e:
        print(f"員工列表錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hr_create_employee(request):
    """
    HR 新增員工

    URL: POST /api/hr/employees/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        data = request.data

        # 必填欄位
        required_fields = ['employee_id', 'username', 'password']
        for field in required_fields:
            if not data.get(field):
                return validation_error_response(f"缺少必填欄位：{field}")

        # 檢查員工編號是否重複
        if Employees.objects.filter(employee_id=data['employee_id']).exists():
            return error_response(
                "員工編號已存在",
                code="DUPLICATE_EMPLOYEE_ID",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        # 建立員工
        employee = Employees(
            employee_id=data['employee_id'],
            username=data['username'],
            email=data.get('email'),
            phone=data.get('phone'),
            address=data.get('address'),
            role=data.get('role', 'employee'),
        )
        employee.set_password(data['password'])

        # 設定部門
        if data.get('department'):
            from .models import Departments
            try:
                department = Departments.objects.get(id=data['department'])
                employee.department = department
            except Departments.DoesNotExist:
                pass

        employee.save()

        # 建立員工-公司關聯
        if data.get('company_id') and data.get('hire_date'):
            EmpCompanyRel.objects.create(
                employee_id=employee,
                company_id_id=data['company_id'],
                employment_status=True,
                hire_date=data['hire_date'],
                direct_manager_id=data.get('direct_manager'),
            )

        return success_response(
            message="員工建立成功",
            data={'employee_id': employee.employee_id},
            status_code=status.HTTP_201_CREATED
        )

    except Exception as e:
        print(f"建立員工錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("建立失敗，請稍後再試")


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def hr_update_employee(request, employee_id):
    """
    HR 更新員工資料

    URL: PATCH /api/hr/employees/{employee_id}/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        # 取得員工
        try:
            employee = Employees.objects.get(employee_id=employee_id)
        except Employees.DoesNotExist:
            return not_found_response("員工不存在")

        data = request.data

        # 更新允許的欄位
        allowed_fields = ['username', 'email', 'phone', 'address', 'role', 'is_active']
        for field in allowed_fields:
            if field in data:
                setattr(employee, field, data[field])

        # 更新部門
        if 'department' in data:
            if data['department']:
                from .models import Departments
                try:
                    department = Departments.objects.get(id=data['department'])
                    employee.department = department
                except Departments.DoesNotExist:
                    pass
            else:
                employee.department = None

        # 更新密碼
        if data.get('password'):
            employee.set_password(data['password'])

        employee.save()

        return success_response(
            message="員工資料已更新",
            data={'employee_id': employee.employee_id}
        )

    except Exception as e:
        print(f"更新員工錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("更新失敗，請稍後再試")


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def hr_assign_manager(request, employee_id):
    """
    HR 指派主管

    URL: PATCH /api/hr/employees/{employee_id}/assign-manager/
    請求參數：
    - manager_id: 主管員工編號
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        manager_id = request.data.get('manager_id')
        if not manager_id:
            return validation_error_response("請提供主管員工編號")

        # 取得員工
        try:
            employee = Employees.objects.get(employee_id=employee_id)
        except Employees.DoesNotExist:
            return not_found_response("員工不存在")

        # 取得主管
        try:
            manager = Employees.objects.get(employee_id=manager_id)
        except Employees.DoesNotExist:
            return not_found_response("主管不存在")

        # 更新 EmpCompanyRel
        relations = EmpCompanyRel.objects.filter(
            employee_id=employee,
            employment_status=True
        )

        if not relations.exists():
            return error_response(
                "員工尚未建立公司關聯",
                code="NO_RELATION",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        for rel in relations:
            rel.direct_manager = manager
            rel.save()

        return success_response(
            message="主管指派成功",
            data={
                'employee_id': employee_id,
                'manager_id': manager_id,
                'manager_name': manager.username
            }
        )

    except Exception as e:
        print(f"指派主管錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("指派失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def hr_batch_set_leave_balances(request):
    """
    HR 批次設定假別額度

    URL: POST /api/hr/leave-balances/batch-set/
    請求參數：
    - employee_ids: 員工編號陣列
    - year: 年度
    - leave_type: 假別
    - total_hours: 總額度
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        employee_ids = request.data.get('employee_ids', [])
        year = request.data.get('year', datetime.now().year)
        leave_type = request.data.get('leave_type')
        total_hours = request.data.get('total_hours')

        if not employee_ids or not leave_type or total_hours is None:
            return validation_error_response("缺少必要參數")

        # 批次處理
        processed = []
        failed = []

        for emp_id in employee_ids:
            try:
                employee = Employees.objects.get(employee_id=emp_id)

                # 更新或建立額度
                balance, created = LeaveBalances.objects.update_or_create(
                    employee_id=employee,
                    year=year,
                    leave_type=leave_type,
                    defaults={
                        'total_hours': Decimal(str(total_hours)),
                    }
                )
                # 重新計算剩餘時數
                balance.remaining_hours = balance.total_hours - balance.used_hours
                balance.save()

                processed.append(emp_id)

            except Employees.DoesNotExist:
                failed.append({
                    'employee_id': emp_id,
                    'reason': '員工不存在'
                })

        return success_response(
            message=f"批次設定完成：成功 {len(processed)} 筆，失敗 {len(failed)} 筆",
            data={
                'processed': processed,
                'failed': failed,
            }
        )

    except Exception as e:
        print(f"批次設定假別額度錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("設定失敗，請稍後再試")


# =====================================================
# Phase 3 新增：部門管理 API
# =====================================================

from .models import Departments


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def department_list(request):
    """
    部門列表

    URL: GET /api/hr/departments/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限存取此功能")

        company_id = request.query_params.get('company_id')

        queryset = Departments.objects.filter(is_active=True)

        if company_id:
            queryset = queryset.filter(company_id=company_id)

        serializer = DepartmentsSerializer(queryset, many=True)

        return success_response(
            message="查詢成功",
            data={
                'count': queryset.count(),
                'departments': serializer.data,
            }
        )

    except Exception as e:
        print(f"部門列表錯誤: {str(e)}")
        return server_error_response("查詢失敗，請稍後再試")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def department_create(request):
    """
    建立部門

    URL: POST /api/hr/departments/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        data = request.data

        if not data.get('name') or not data.get('company_id'):
            return validation_error_response("缺少必要參數（name, company_id）")

        # 檢查是否重複
        if Departments.objects.filter(
            name=data['name'],
            company_id=data['company_id']
        ).exists():
            return error_response(
                "部門名稱已存在",
                code="DUPLICATE_DEPARTMENT",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        department = Departments.objects.create(
            name=data['name'],
            company_id_id=data['company_id'],
            manager_id=data.get('manager'),
            parent_department_id=data.get('parent_department'),
            description=data.get('description', ''),
        )

        serializer = DepartmentsSerializer(department)

        return success_response(
            message="部門建立成功",
            data=serializer.data,
            status_code=status.HTTP_201_CREATED
        )

    except Exception as e:
        print(f"建立部門錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("建立失敗，請稍後再試")


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def department_update(request, department_id):
    """
    更新部門

    URL: PATCH /api/hr/departments/{department_id}/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        try:
            department = Departments.objects.get(id=department_id)
        except Departments.DoesNotExist:
            return not_found_response("部門不存在")

        data = request.data

        # 更新欄位
        if 'name' in data:
            department.name = data['name']
        if 'manager' in data:
            department.manager_id = data['manager']
        if 'parent_department' in data:
            department.parent_department_id = data['parent_department']
        if 'description' in data:
            department.description = data['description']
        if 'is_active' in data:
            department.is_active = data['is_active']

        department.save()

        serializer = DepartmentsSerializer(department)

        return success_response(
            message="部門已更新",
            data=serializer.data
        )

    except Exception as e:
        print(f"更新部門錯誤: {str(e)}")
        return server_error_response("更新失敗，請稍後再試")


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def department_delete(request, department_id):
    """
    刪除部門（軟刪除）

    URL: DELETE /api/hr/departments/{department_id}/
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_hr_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        try:
            department = Departments.objects.get(id=department_id)
        except Departments.DoesNotExist:
            return not_found_response("部門不存在")

        # 軟刪除
        department.is_active = False
        department.save()

        return success_response(message="部門已刪除")

    except Exception as e:
        print(f"刪除部門錯誤: {str(e)}")
        return server_error_response("刪除失敗，請稍後再試")


# =====================================================
# Phase 3 新增：資料匯出 API
# =====================================================

import csv
import io
from django.http import HttpResponse


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_attendance(request):
    """
    匯出出勤記錄

    URL: POST /api/export/attendance/
    請求參數：
    - date_from: 開始日期
    - date_to: 結束日期
    - format: 格式（csv/xlsx）
    - employee_ids: 員工編號陣列（選填，HR 專用）
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_manager_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        export_format = request.data.get('format', 'csv')
        employee_ids = request.data.get('employee_ids', [])

        if not date_from or not date_to:
            return validation_error_response("請提供日期範圍")

        # 解析日期
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()

        # 取得員工（根據權限）
        if _check_hr_permission(user) and employee_ids:
            # HR 可指定員工
            relations = EmpCompanyRel.objects.filter(
                employee_id__in=employee_ids,
                employment_status=True
            )
        else:
            # 主管只能匯出下屬
            subordinates = _get_subordinates(user)
            subordinate_ids = [e.employee_id for e in subordinates]
            relations = EmpCompanyRel.objects.filter(
                employee_id__in=subordinate_ids,
                employment_status=True
            )

        relation_ids = relations.values_list('id', flat=True)

        # 查詢記錄
        records = AttendanceRecords.objects.filter(
            relation_id__in=relation_ids,
            date__gte=start_date,
            date__lte=end_date
        ).select_related('relation_id__employee_id').order_by('date', 'relation_id')

        if export_format == 'csv':
            return _export_attendance_csv(records)
        elif export_format == 'xlsx':
            return _export_attendance_xlsx(records)
        else:
            return validation_error_response("不支援的匯出格式")

    except Exception as e:
        print(f"匯出出勤錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return server_error_response("匯出失敗，請稍後再試")


def _export_attendance_csv(records):
    """匯出出勤記錄為 CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # 標題列
    writer.writerow([
        '日期', '員工編號', '員工姓名', '上班時間', '下班時間',
        '工時', '是否遲到', '遲到分鐘', '是否早退', '早退分鐘', '是否補打卡'
    ])

    # 資料列
    for record in records:
        writer.writerow([
            str(record.date),
            record.relation_id.employee_id.employee_id,
            record.relation_id.employee_id.username,
            str(record.checkin_time) if record.checkin_time else '',
            str(record.checkout_time) if record.checkout_time else '',
            str(record.work_hours),
            '是' if record.is_late else '否',
            record.late_minutes,
            '是' if record.is_early_leave else '否',
            record.early_leave_minutes,
            '是' if record.is_makeup else '否',
        ])

    output.seek(0)

    response = HttpResponse(
        output.read().encode('utf-8-sig'),
        content_type='text/csv; charset=utf-8-sig'
    )
    response['Content-Disposition'] = 'attachment; filename="attendance_export.csv"'

    return response


def _export_attendance_xlsx(records):
    """匯出出勤記錄為 Excel"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出勤記錄"

        # 標題列
        headers = [
            '日期', '員工編號', '員工姓名', '上班時間', '下班時間',
            '工時', '是否遲到', '遲到分鐘', '是否早退', '早退分鐘', '是否補打卡'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 資料列
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=str(record.date))
            ws.cell(row=row, column=2, value=record.relation_id.employee_id.employee_id)
            ws.cell(row=row, column=3, value=record.relation_id.employee_id.username)
            ws.cell(row=row, column=4, value=str(record.checkin_time) if record.checkin_time else '')
            ws.cell(row=row, column=5, value=str(record.checkout_time) if record.checkout_time else '')
            ws.cell(row=row, column=6, value=float(record.work_hours))
            ws.cell(row=row, column=7, value='是' if record.is_late else '否')
            ws.cell(row=row, column=8, value=record.late_minutes)
            ws.cell(row=row, column=9, value='是' if record.is_early_leave else '否')
            ws.cell(row=row, column=10, value=record.early_leave_minutes)
            ws.cell(row=row, column=11, value='是' if record.is_makeup else '否')

        # 調整列寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance_export.xlsx"'

        return response

    except ImportError:
        return error_response(
            "伺服器未安裝 openpyxl，請使用 CSV 格式",
            code="OPENPYXL_NOT_INSTALLED",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_leave(request):
    """
    匯出請假記錄

    URL: POST /api/export/leave/
    請求參數：
    - date_from: 開始日期
    - date_to: 結束日期
    - format: 格式（csv/xlsx）
    """
    try:
        user = request.user

        # 權限檢查
        if not _check_manager_permission(user):
            return forbidden_response("您沒有權限執行此操作")

        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        export_format = request.data.get('format', 'csv')

        if not date_from or not date_to:
            return validation_error_response("請提供日期範圍")

        # 解析日期
        start_date = datetime.strptime(date_from, '%Y-%m-%d')
        end_date = datetime.strptime(date_to, '%Y-%m-%d')

        # 取得員工
        subordinates = _get_subordinates(user)
        subordinate_ids = [e.employee_id for e in subordinates]
        relations = EmpCompanyRel.objects.filter(
            employee_id__in=subordinate_ids,
            employment_status=True
        )
        relation_ids = relations.values_list('id', flat=True)

        # 查詢記錄
        records = LeaveRecords.objects.filter(
            relation_id__in=relation_ids,
            start_time__gte=start_date,
            start_time__lte=end_date
        ).select_related('relation_id__employee_id').order_by('start_time')

        if export_format == 'csv':
            return _export_leave_csv(records)
        else:
            return validation_error_response("目前僅支援 CSV 格式")

    except Exception as e:
        print(f"匯出請假錯誤: {str(e)}")
        return server_error_response("匯出失敗，請稍後再試")


def _export_leave_csv(records):
    """匯出請假記錄為 CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # 標題列
    writer.writerow([
        '員工編號', '員工姓名', '假別', '開始時間', '結束時間',
        '請假時數', '狀態', '請假原因'
    ])

    # 資料列
    for record in records:
        writer.writerow([
            record.relation_id.employee_id.employee_id,
            record.relation_id.employee_id.username,
            record.get_leave_type_display(),
            str(record.start_time),
            str(record.end_time),
            str(record.leave_hours),
            record.get_status_display(),
            record.leave_reason or '',
        ])

    output.seek(0)

    response = HttpResponse(
        output.read().encode('utf-8-sig'),
        content_type='text/csv; charset=utf-8-sig'
    )
    response['Content-Disposition'] = 'attachment; filename="leave_export.csv"'

    return response
