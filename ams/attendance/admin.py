from django.contrib import admin
from django.http import HttpResponse
from django import forms
from django.shortcuts import render
from django.urls import path, reverse
from django.utils.html import format_html
import openpyxl
import qrcode
from io import BytesIO
import json
from .models import (
    Employees, Companies, EmpCompanyRel, LeaveRecords,
    AttendanceRecords, ApprovalRecords, LeaveBalances,
    ManagerialRelationship, ApprovalPolicy,
    # Phase 1 新增
    WorkSchedule, MakeupClockRequest, MakeupClockApproval, MakeupClockQuota,
    # Phase 2 新增
    OvertimeRecords, OvertimeApproval, Notifications,
    # Phase 3 新增
    Departments
)

class DateRangeForm(forms.Form):
    start_date = forms.DateField(label="起始日期", widget=forms.DateInput(attrs={'type': 'date'}))
    end_date = forms.DateField(label="結束日期", widget=forms.DateInput(attrs={'type': 'date'}))


@admin.register(Employees)  # 這樣 Django Admin 就能識別 Employees 模型
class EmployeesAdmin(admin.ModelAdmin):
    # Phase 3 更新：顯示角色和部門欄位
    list_display = ('employee_id', 'username', 'role', 'department', 'phone', 'is_active')
    list_filter = ('role', 'department', 'is_active')
    search_fields = ['employee_id', 'username', 'email']

    fieldsets = (
        ('基本資訊', {
            'fields': ('employee_id', 'username', 'email', 'phone', 'address')
        }),
        ('角色與部門', {
            'fields': ('role', 'department')
        }),
        ('帳戶狀態', {
            'fields': ('is_active', 'is_staff', 'is_superuser')
        }),
    )


@admin.register(Companies)
class CompaniesAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'address', 'latitude', 'longitude', 'radius', 'qr_code_download')

    def get_urls(self):
        """新增自訂 URL 路由"""
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:company_id>/qrcode/',
                self.admin_site.admin_view(self.download_qrcode),
                name='company-qrcode-download',
            ),
        ]
        return custom_urls + urls

    def qr_code_download(self, obj):
        """在列表中顯示下載 QR Code 連結"""
        url = reverse('admin:company-qrcode-download', args=[obj.id])
        return format_html(
            '<a href="{}" class="button" style="padding: 5px 10px; background: #417690; color: white; '
            'text-decoration: none; border-radius: 4px;">📥 下載 QR Code</a>',
            url
        )
    qr_code_download.short_description = 'QR Code'
    qr_code_download.allow_tags = True

    def download_qrcode(self, request, company_id):
        """產生並下載 QR Code 圖檔"""
        try:
            company = Companies.objects.get(id=company_id)
        except Companies.DoesNotExist:
            return HttpResponse('公司不存在', status=404)

        # QR Code 內容：JSON 格式包含公司資訊
        qr_data = json.dumps({
            'type': 'attendance_clock',
            'company_id': company.id,
            'company_name': company.name,
            'latitude': company.latitude,
            'longitude': company.longitude,
            'radius': company.radius,
        }, ensure_ascii=False)

        # 產生 QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        # 產生圖片
        img = qr.make_image(fill_color="black", back_color="white")

        # 轉換為 bytes
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        # 回傳圖檔
        response = HttpResponse(buffer.getvalue(), content_type='image/png')
        filename = f"qrcode_{company.name}_{company.id}.png"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@admin.register(EmpCompanyRel)
class CEmpCompanyRelAdmin(admin.ModelAdmin):
    list_display = ('id', 'employee_id_display', 'employee_name_display',
        'company_id_display', 'company_name_display', 'direct_manager', 'employment_status', 'hire_date', 'leave_date'
    )

    def employee_id_display(self, obj):
        return obj.employee_id  # 不加 `.id`，因為它是數值欄位
    employee_id_display.short_description = '員工編號'

    def employee_name_display(self, obj):
        return obj.employee_id.username
    employee_name_display.short_description = '員工姓名'

    def company_id_display(self, obj):
        return obj.company_id.id
    company_id_display.short_description = '公司編號'

    def company_name_display(self, obj):
        return obj.company_id.name
    company_name_display.short_description = '公司名稱'

    search_fields = ['employee_id__employee_id',
                    'employee_id__username']


@admin.register(LeaveRecords)
class LeaveRecordsAdmin(admin.ModelAdmin):
    list_display = ('relation_id_display', 'employee_id_display', 'employee_name_display',
        'company_name_display', 'start_time', 'end_time', 'leave_hours', 'leave_reason'
    )
    def relation_id_display(self, obj):
        return obj.relation_id.id
    relation_id_display.short_description = '關聯編號'

    def employee_id_display(self, obj):
        return obj.relation_id.employee_id
    employee_id_display.short_description = '員工編號'

    def employee_name_display(self, obj):
        return obj.relation_id.employee_id.username
    employee_name_display.short_description = '員工姓名'

    def company_name_display(self, obj):
        try:
            return obj.relation_id.company_id.name
        except Exception as e:
            return f"錯誤: {str(e)}"
    company_name_display.short_description = '公司名稱'


    actions = ['export_by_date_range']
    search_fields = ['relation_id__employee_id__employee_id',
                    'relation_id__employee_id__username']


    @admin.action(description='匯出指定日期範圍的資料')
    def export_by_date_range(self, request, queryset):
        if 'apply' in request.POST:
            form = DateRangeForm(request.POST)
            if form.is_valid():
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']

                filtered = queryset.filter(start_time__date__range=(start_date, end_date)).order_by(
                    'relation_id__employee_id__employee_id',
                    'start_time'
                )

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "請假紀錄"

                headers = ['員工編號', '員工姓名', '關聯編號', '日期', '請假開始時間', '請假結束時間', '請假總時數', '請假原因']
                ws.append(headers)

                for record in filtered:
                    emp = record.relation_id.employee_id
                    ws.append([
                        emp.employee_id,
                        emp.username,
                        record.relation_id.id,
                        record.start_time.strftime('%H:%M:%S') if record.start_time else '',
                        record.end_time.strftime('%H:%M:%S') if record.end_time else '',
                        record.leave_hours or '',
                        record.leave_reason or '',
                    ])

                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename=attendance_filtered.xlsx'
                wb.save(response)
                return response
        else:
            form = DateRangeForm()

        return render(request, 'admin/date_range_form.html', {
            'form': form,
            'queryset': queryset,
            'action': 'export_by_date_range'
        })

@admin.register(ApprovalRecords)
class ApprovalRecordsAdmin(admin.ModelAdmin):
    list_display = ('id', 'leave_id', 'approver_id', 'approval_level', 'status', 'created_at', 'approved_at')
    list_filter = ('status', 'approval_level')
    search_fields = ['leave_id__relation_id__employee_id__username', 'approver_id__username']

    def save_model(self, request, obj, form, change):
        """覆寫 save_model，當審批狀態改變時同步更新請假記錄"""
        super().save_model(request, obj, form, change)

        # 如果審批通過
        if obj.status == 'approved':
            leave = obj.leave_id
            # 檢查是否還有其他待審批的記錄
            pending_count = ApprovalRecords.objects.filter(
                leave_id=leave,
                status='pending'
            ).exclude(id=obj.id).count()

            # 如果沒有待審批，更新請假狀態並扣除額度
            if pending_count == 0 and leave.status == 'pending':
                leave.status = 'approved'
                leave.save()

                # 扣除假別額度
                from attendance.models import LeaveBalances
                employee = leave.relation_id.employee_id
                year = leave.start_time.year
                balance = LeaveBalances.objects.filter(
                    employee_id=employee,
                    year=year,
                    leave_type=leave.leave_type
                ).first()
                if balance:
                    balance.used_hours += leave.leave_hours
                    balance.save()

        # 如果審批拒絕
        elif obj.status == 'rejected':
            leave = obj.leave_id
            leave.status = 'rejected'
            leave.save()

@admin.register(LeaveBalances)
class LeaveBalancesAdmin(admin.ModelAdmin):
    list_display = ('id', 'employee_name_display', 'year', 'leave_type_display', 'total_hours', 'used_hours', 'remaining_hours')
    list_filter = ('year', 'leave_type')
    search_fields = ['employee_id__username', 'employee_id__employee_id']

    def employee_name_display(self, obj):
        return f"{obj.employee_id.username} ({obj.employee_id.employee_id})"
    employee_name_display.short_description = '員工'

    def leave_type_display(self, obj):
        return obj.get_leave_type_display()
    leave_type_display.short_description = '假別'

@admin.register(AttendanceRecords)
class AttendanceRecordsAdmin(admin.ModelAdmin):
    list_display = ('relation_id_display', 'employee_code_display', 'user_name',
        'company_name', 'date', 'checkin_time', 'checkout_time', 'checkin_location', 'checkout_location', 'work_hours'
    )

    def user_name(self, obj):
        return obj.relation_id.employee_id.username
    user_name.short_description = '員工姓名'
    
    def employee_code_display(self, obj):
        return obj.relation_id.employee_id.employee_id  # 這是 Employees 表的主鍵
    employee_code_display.short_description = '員工編號'

    def relation_id_display(self, obj):
        return obj.relation_id.id
    relation_id_display.short_description = '關聯編號'

    def company_name(self, obj):
        try:
            return obj.relation_id.company_id.name
        except Exception as e:
            return f"錯誤: {str(e)}"
    company_name.short_description = '公司名稱'



    actions = ['export_attendance_and_leave']
    search_fields = ['relation_id__employee_id__employee_id',
                    'relation_id__employee_id__username']


    @admin.action(description="匯出出缺勤與請假紀錄")
    def export_attendance_and_leave(self, request, queryset):
        if 'apply' in request.POST:
            form = DateRangeForm(request.POST)
            if form.is_valid():
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']

                # === 匯出 Excel 初始化 ===
                wb = openpyxl.Workbook()

                # === 出缺勤紀錄 Sheet ===
                ws1 = wb.active
                ws1.title = "出缺勤紀錄"

                headers1 = ['員工編號', '員工姓名', '關聯編號', '日期', '簽到時間', '簽退時間', '簽到地點', '簽退地點', '工作時數']
                ws1.append(headers1)

                attendance_records = AttendanceRecords.objects.filter(
                    date__range=(start_date, end_date)
                ).order_by('relation_id__employee_id__employee_id', 'date')

                for record in attendance_records:
                    emp = record.relation_id.employee_id
                    ws1.append([
                        emp.employee_id,
                        emp.username,
                        record.relation_id.id,
                        record.date.strftime('%Y-%m-%d') if record.date else '',
                        record.checkin_time.strftime('%H:%M:%S') if record.checkin_time else '',
                        record.checkout_time.strftime('%H:%M:%S') if record.checkout_time else '',
                        record.checkin_location or '',
                        record.checkout_location or '',
                        record.work_hours or '',
                    ])

                # === 請假紀錄 Sheet ===
                ws2 = wb.create_sheet(title="請假紀錄")

                headers2 = ['員工編號', '員工姓名', '關聯編號', '請假日期', '開始時間', '結束時間', '請假時數', '請假原因']
                ws2.append(headers2)

                leave_records = LeaveRecords.objects.filter(
                    start_time__date__range=(start_date, end_date)
                ).order_by('relation_id__employee_id__employee_id', 'start_time')

                for record in leave_records:
                    emp = record.relation_id.employee_id
                    leave_date = record.start_time.date() if record.start_time else ''
                    ws2.append([
                        emp.employee_id,
                        emp.username,
                        record.relation_id.id,
                        leave_date.strftime('%Y-%m-%d') if leave_date else '',
                        record.start_time.strftime('%H:%M:%S') if record.start_time else '',
                        record.end_time.strftime('%H:%M:%S') if record.end_time else '',
                        record.leave_hours or '',
                        record.leave_reason or '',
                    ])

                # === 回傳 Excel 檔案 ===
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename=attendance_leave_summary.xlsx'
                wb.save(response)
                return response
        else:
            form = DateRangeForm()

        return render(request, 'admin/date_range_form.html', {
            'form': form,
            'queryset': queryset,
            'action': 'export_attendance_and_leave'
        })



# ========== 審批管理系統 Admin ==========

@admin.register(ManagerialRelationship)
class ManagerialRelationshipAdmin(admin.ModelAdmin):
    list_display = ('id', 'employee_display', 'manager_display', 'company_id', 'effective_date', 'end_date', 'created_by')
    list_filter = ('company_id', 'effective_date')
    search_fields = ['employee_id__employee_id', 'employee_id__username', 'manager_id__employee_id', 'manager_id__username']
    date_hierarchy = 'effective_date'

    def employee_display(self, obj):
        return f"{obj.employee_id.username} ({obj.employee_id.employee_id})"
    employee_display.short_description = '員工'

    def manager_display(self, obj):
        return f"{obj.manager_id.username} ({obj.manager_id.employee_id})"
    manager_display.short_description = '直屬主管'

    def has_add_permission(self, request):
        """只有 HR 和總經理可以新增"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def has_change_permission(self, request, obj=None):
        """只有 HR 和總經理可以修改"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def has_delete_permission(self, request, obj=None):
        """只有 HR 和總經理可以刪除"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def save_model(self, request, obj, form, change):
        """記錄建立者"""
        if not change:  # 新增時
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

        # 同步更新 EmpCompanyRel 的 direct_manager
        try:
            rel = EmpCompanyRel.objects.filter(
                employee_id=obj.employee_id,
                company_id=obj.company_id
            ).first()
            if rel:
                rel.direct_manager = obj.manager_id
                rel.save()
        except Exception as e:
            print(f"同步主管關係錯誤: {str(e)}")


@admin.register(ApprovalPolicy)
class ApprovalPolicyAdmin(admin.ModelAdmin):
    list_display = ('id', 'policy_name', 'company_id', 'min_days', 'max_days', 'is_active', 'created_at')
    list_filter = ('is_active', 'company_id')
    search_fields = ['policy_name']

    fieldsets = (
        ('基本資訊', {
            'fields': ('policy_name', 'company_id', 'is_active')
        }),
        ('天數範圍', {
            'fields': ('min_days', 'max_days'),
            'description': '定義此政策適用的請假天數範圍'
        }),
        ('審批層級', {
            'fields': ('approval_levels',),
            'description': '''
            JSON 格式範例：
            [
                {"level": 1, "role": "manager", "description": "直屬主管"},
                {"level": 2, "role": "hr", "description": "人資部門"},
                {"level": 3, "role": "ceo", "description": "總經理"}
            ]

            role 可選值：
            - manager: 直屬主管（自動取得）
            - hr: 人資部門（employee_id 以 HR 開頭）
            - ceo: 總經理（employee_id 以 CEO 開頭）
            - custom: 自訂（需要額外處理）
            '''
        }),
    )

    def has_add_permission(self, request):
        """只有 HR 和總經理可以新增"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def has_change_permission(self, request, obj=None):
        """只有 HR 和總經理可以修改"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def has_delete_permission(self, request, obj=None):
        """只有 HR 和總經理可以刪除"""
        if request.user.is_superuser:
            return True
        emp_id = getattr(request.user, 'employee_id', '')
        return emp_id.startswith('HR') or emp_id.startswith('CEO')

    def save_model(self, request, obj, form, change):
        """記錄建立者"""
        if not change:  # 新增時
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


# =====================================================
# Phase 1 新增：工時設定與補打卡 Admin
# =====================================================

@admin.register(WorkSchedule)
class WorkScheduleAdmin(admin.ModelAdmin):
    """工時設定管理"""
    list_display = ('id', 'company_id', 'name', 'work_start_time', 'work_end_time',
                    'standard_work_hours', 'grace_period_minutes', 'is_default', 'is_active')
    list_filter = ('company_id', 'is_default', 'is_active')
    search_fields = ['name', 'company_id__name']

    fieldsets = (
        ('基本資訊', {
            'fields': ('company_id', 'name', 'is_default', 'is_active')
        }),
        ('工時設定', {
            'fields': ('work_start_time', 'work_end_time', 'standard_work_hours', 'lunch_break_minutes'),
            'description': '設定標準上下班時間與工時'
        }),
        ('遲到設定', {
            'fields': ('grace_period_minutes',),
            'description': '遲到寬限時間，在此時間內打卡不算遲到'
        }),
    )


@admin.register(MakeupClockRequest)
class MakeupClockRequestAdmin(admin.ModelAdmin):
    """補打卡申請管理"""
    list_display = ('id', 'employee_display', 'date', 'makeup_type', 'status', 'reason_short', 'created_at')
    list_filter = ('status', 'makeup_type', 'date')
    search_fields = ['relation_id__employee_id__username', 'relation_id__employee_id__employee_id', 'reason']
    date_hierarchy = 'date'
    readonly_fields = ('created_at', 'updated_at')

    def employee_display(self, obj):
        return f"{obj.relation_id.employee_id.username} ({obj.relation_id.employee_id.employee_id})"
    employee_display.short_description = '申請人'

    def reason_short(self, obj):
        return obj.reason[:30] + '...' if len(obj.reason) > 30 else obj.reason
    reason_short.short_description = '原因'


@admin.register(MakeupClockApproval)
class MakeupClockApprovalAdmin(admin.ModelAdmin):
    """補打卡審批記錄管理"""
    list_display = ('id', 'request_id', 'approver_display', 'approval_level', 'status', 'approved_at')
    list_filter = ('status', 'approval_level')
    search_fields = ['request_id__relation_id__employee_id__username', 'approver_id__username']
    readonly_fields = ('created_at',)

    def approver_display(self, obj):
        return f"{obj.approver_id.username} ({obj.approver_id.employee_id})"
    approver_display.short_description = '審批人'


@admin.register(MakeupClockQuota)
class MakeupClockQuotaAdmin(admin.ModelAdmin):
    """補打卡額度管理"""
    list_display = ('id', 'employee_display', 'year', 'total_count', 'used_count', 'remaining_count')
    list_filter = ('year',)
    search_fields = ['employee_id__username', 'employee_id__employee_id']

    def employee_display(self, obj):
        return f"{obj.employee_id.username} ({obj.employee_id.employee_id})"
    employee_display.short_description = '員工'


# =====================================================
# Phase 2 新增：加班管理 Admin
# =====================================================

@admin.register(OvertimeRecords)
class OvertimeRecordsAdmin(admin.ModelAdmin):
    """加班記錄管理"""
    list_display = ('id', 'employee_display', 'date', 'start_time', 'end_time',
                    'overtime_hours', 'compensation_type', 'status', 'created_at')
    list_filter = ('status', 'compensation_type', 'date')
    search_fields = ['relation_id__employee_id__username', 'relation_id__employee_id__employee_id']
    readonly_fields = ('created_at', 'updated_at')
    date_hierarchy = 'date'

    def employee_display(self, obj):
        return f"{obj.relation_id.employee_id.username} ({obj.relation_id.employee_id.employee_id})"
    employee_display.short_description = '員工'


@admin.register(OvertimeApproval)
class OvertimeApprovalAdmin(admin.ModelAdmin):
    """加班審批記錄管理"""
    list_display = ('id', 'overtime_id', 'approver_display', 'approval_level', 'status', 'approved_at')
    list_filter = ('status', 'approval_level')
    search_fields = ['overtime_id__relation_id__employee_id__username', 'approver_id__username']
    readonly_fields = ('created_at',)

    def approver_display(self, obj):
        return f"{obj.approver_id.username} ({obj.approver_id.employee_id})"
    approver_display.short_description = '審批人'


# =====================================================
# Phase 2 新增：通知系統 Admin
# =====================================================

@admin.register(Notifications)
class NotificationsAdmin(admin.ModelAdmin):
    """通知記錄管理"""
    list_display = ('id', 'recipient_display', 'notification_type', 'title', 'is_read', 'created_at')
    list_filter = ('notification_type', 'is_read', 'created_at')
    search_fields = ['recipient_id__username', 'title', 'content']
    readonly_fields = ('created_at',)
    date_hierarchy = 'created_at'

    def recipient_display(self, obj):
        return f"{obj.recipient_id.username} ({obj.recipient_id.employee_id})"
    recipient_display.short_description = '接收人'


# =====================================================
# Phase 3 新增：部門管理 Admin
# =====================================================

@admin.register(Departments)
class DepartmentsAdmin(admin.ModelAdmin):
    """部門管理"""
    list_display = ('id', 'name', 'company_id', 'manager_display', 'parent_department', 'employee_count', 'is_active')
    list_filter = ('company_id', 'is_active')
    search_fields = ['name', 'company_id__name', 'manager__username']
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        ('基本資訊', {
            'fields': ('name', 'company_id', 'description', 'is_active')
        }),
        ('組織架構', {
            'fields': ('manager', 'parent_department')
        }),
        ('時間戳記', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def manager_display(self, obj):
        if obj.manager:
            return f"{obj.manager.username} ({obj.manager.employee_id})"
        return '-'
    manager_display.short_description = '部門主管'

    def employee_count(self, obj):
        return obj.get_employee_count()
    employee_count.short_description = '員工數'
